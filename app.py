import csv
import hashlib
import io
import json
import os
import secrets
import shutil
import sqlite3
import tempfile
from datetime import datetime, timedelta, timezone
from contextlib import contextmanager
from functools import wraps
import click
from flask import Flask, render_template, render_template_string, jsonify, request, send_file, redirect, url_for, flash, session, has_request_context
from flask_login import LoginManager, UserMixin, login_user, logout_user, login_required, current_user
from jinja2 import TemplateNotFound
from werkzeug.security import generate_password_hash, check_password_hash
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address

import reservas as reservas_mod
import relatorios
import agendamento_utils
import backup_utils
import email_utils
import excel_import_utils
import system_utils
import usuarios as usuarios_mod
from user_utils import normalizar_supervisor_id, sugerir_username_por_nome, valor_ativo
from migrations import adicionar_coluna_se_ausente, executar_migration


# ========================================
# CONFIGURACAO
# ========================================

app = Flask(__name__)
try:
    from dotenv import load_dotenv as _ld
    _ld(dotenv_path='/home/victroid/mapa_sala/.env')
except ImportError:
    pass

_secret_key = os.environ.get('SECRET_KEY')
if not _secret_key:
    raise RuntimeError(
        "SECRET_KEY não definida. Crie o arquivo .env com SECRET_KEY=<chave> "
        "ou defina a variável de ambiente antes de iniciar o app."
    )

app.secret_key = _secret_key
app.permanent_session_lifetime = timedelta(minutes=30)
app.config['SESSION_REFRESH_EACH_REQUEST'] = True
DB_PATH = os.environ.get(
    'DB_PATH',
    os.path.join(os.path.dirname(os.path.abspath(__file__)), 'mapa_salas.db')
)

VERSAO = '2026-07-08-v41'
EMAIL_BASE_URL = os.environ.get('EMAIL_BASE_URL', '').rstrip('/')
EMAIL_FROM = os.environ.get('EMAIL_FROM', os.environ.get('SMTP_USER', ''))
SMTP_HOST = os.environ.get('SMTP_HOST', '')
SMTP_PORT = int(os.environ.get('SMTP_PORT', '587'))
SMTP_USER = os.environ.get('SMTP_USER', '')
SMTP_PASSWORD = os.environ.get('SMTP_PASSWORD', '')
SMTP_TLS = os.environ.get('SMTP_TLS', '1').strip().lower() not in ('0', 'false', 'nao', 'não')
BACKUP_DIR = os.environ.get(
    'BACKUP_DIR',
    os.path.join(os.path.dirname(os.path.abspath(__file__)), 'backups')
)
BACKUP_RETENTION_DAYS = int(os.environ.get('BACKUP_RETENTION_DAYS', '30'))

login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'
login_manager.login_message = 'Faça login para acessar o sistema.'

limiter = Limiter(
    get_remote_address,
    app=app,
    default_limits=[],
    storage_uri='memory://'
)

PAPEIS_VALIDOS = ('coordenador', 'recepcao', 'professor', 'aluno', 'somente_leitura')
STATUS_SOLICITACAO_VAGA = {
    'pendente': 'Pendente',
    'em_analise': 'Em análise',
    'atendida_parcial': 'Atendida parcialmente',
    'atendida': 'Atendida',
    'recusada': 'Recusada',
}


def preparar_solicitacao_vaga(row):
    item = dict(row)
    item['status_label'] = STATUS_SOLICITACAO_VAGA.get(item.get('status'), item.get('status'))
    item['vagas_paciente'] = int(item.get('vagas_paciente') or 0)
    item['vagas_triagem'] = int(item.get('vagas_triagem') or 0)
    item['vagas_paciente_liberadas'] = int(item.get('vagas_paciente_liberadas') or 0)
    item['vagas_triagem_liberadas'] = int(item.get('vagas_triagem_liberadas') or 0)
    item['vagas_paciente_faltantes'] = max(0, item['vagas_paciente'] - item['vagas_paciente_liberadas'])
    item['vagas_triagem_faltantes'] = max(0, item['vagas_triagem'] - item['vagas_triagem_liberadas'])
    return item


# ========================================
# MODELOS E PERMISSOES
# ========================================

def gerar_csrf_token():
    if '_csrf_token' not in session:
        session['_csrf_token'] = secrets.token_urlsafe(32)
    return session['_csrf_token']


def rota_inicial_por_papel(role):
    if role == 'professor':
        return '/minha-supervisao'
    if role == 'aluno':
        return '/meus-agendamentos'
    return '/mapa'


def itens_menu_por_papel(role):
    menus = {
        'coordenador': [
            ('Rotina', [
                ('Painel da Recepção', '/painel', 'layout-dashboard'),
                ('Afazeres', '/afazeres', 'list-checks'),
                ('Reservas', '/reservas', 'bell'),
                ('Horários abertos', '/horarios-abertos', 'calendar-plus'),
                ('Coordenação', '/coordenacao', 'messages-square'),
            ]),
            ('Gestão', [
                ('Painel da Coordenação', '/painel-coordenacao', 'bar-chart-3'),
                ('Relatório semanal', '/relatorio-semanal', 'file-bar-chart'),
                ('Usuários', '/usuarios', 'users'),
            ]),
            ('Saídas', [
                ('Imprimir', '/imprimir', 'printer'),
            ]),
            ('Sistema', [
                ('Saúde e manutenção', '/saude', 'activity'),
            ]),
            ('Conta', [
                ('Perfil', '/perfil', 'user-round'),
                ('Ajuda', '/ajuda/coordenacao', 'book-open'),
                ('Informações', '/informacoes', 'info'),
            ]),
        ],
        'recepcao': [
            ('Rotina', [
                ('Painel', '/painel', 'layout-dashboard'),
                ('Afazeres', '/afazeres', 'list-checks'),
                ('Reservas', '/reservas', 'bell'),
                ('Horários abertos', '/horarios-abertos', 'calendar-plus'),
            ]),
            ('Saídas', [
                ('Imprimir', '/imprimir', 'printer'),
            ]),
            ('Conta', [
                ('Perfil', '/perfil', 'user-round'),
                ('Ajuda', '/ajuda/recepcao', 'book-open'),
                ('Informações', '/informacoes', 'info'),
            ]),
        ],
        'professor': [
            ('Supervisão', [
                ('Minha Supervisão', '/minha-supervisao', 'users-round'),
                ('Pedidos de vagas', '/minha-supervisao#solicitacoes-vagas', 'clipboard-plus'),
            ]),
            ('Conta', [
                ('Perfil', '/perfil', 'user-round'),
                ('Ajuda', '/ajuda/professor', 'book-open'),
                ('Informações', '/informacoes', 'info'),
            ]),
        ],
        'aluno': [
            ('Minha rotina', [
                ('Meus Agendamentos', '/meus-agendamentos', 'calendar-days'),
                ('Reservar sala', '/reservas#reserva-sala', 'door-open'),
                ('Reservar instrumentos', '/reservas#reserva-instrumentos', 'clipboard-list'),
                ('Horário com Coordenação', '/coordenacao', 'messages-square'),
            ]),
            ('Conta', [
                ('Perfil', '/perfil', 'user-round'),
                ('Ajuda', '/ajuda/aluno', 'book-open'),
                ('Informações', '/informacoes', 'info'),
            ]),
        ],
        'somente_leitura': [
            ('Rotina', [
                ('Painel da Recepção', '/painel', 'layout-dashboard'),
                ('Afazeres', '/afazeres', 'list-checks'),
                ('Reservas', '/reservas', 'bell'),
                ('Horários abertos', '/horarios-abertos', 'calendar-plus'),
                ('Coordenação', '/coordenacao', 'messages-square'),
            ]),
            ('Gestão', [
                ('Painel da Coordenação', '/painel-coordenacao', 'bar-chart-3'),
                ('Relatório semanal', '/relatorio-semanal', 'file-bar-chart'),
                ('Usuários', '/usuarios', 'users'),
            ]),
            ('Saídas', [
                ('Imprimir', '/imprimir', 'printer'),
            ]),
            ('Sistema', [
                ('Saúde e manutenção', '/saude', 'activity'),
            ]),
            ('Conta', [
                ('Perfil', '/perfil', 'user-round'),
                ('Ajuda', '/ajuda/coordenacao', 'book-open'),
                ('Informações', '/informacoes', 'info'),
            ]),
        ],
    }
    return menus.get(role, [])


@app.context_processor
def injetar_csrf_token():
    role = current_user.role if current_user.is_authenticated else ''
    return {
        'csrf_token': gerar_csrf_token,
        'versao': VERSAO,
        'papeis_label': PAPEIS_LABEL,
        'reservas_pendentes_count': contar_reservas_pendentes(),
        'nav_home_url': rota_inicial_por_papel(role),
        'nav_items': itens_menu_por_papel(role),
    }


@app.before_request
def proteger_csrf():
    if request.method not in ('POST', 'PUT', 'PATCH', 'DELETE'):
        return None

    token_salvo = session.get('_csrf_token')
    token_enviado = request.headers.get('X-CSRFToken') or request.form.get('csrf_token')
    if token_salvo and token_enviado and secrets.compare_digest(token_salvo, token_enviado):
        return None

    if request.path.startswith('/api/'):
        return jsonify({'erro': 'Sua sessão expirou. Recarregue a página e tente novamente.'}), 400

    flash('Sua sessão expirou. Recarregue a página e tente novamente.', 'error')
    return redirect(url_for('login'))


@app.before_request
def encerrar_sessao_inativa():
    if not current_user.is_authenticated:
        return None

    agora = datetime.now(timezone.utc)
    ultimo_uso_txt = session.get('ultimo_uso')
    if ultimo_uso_txt:
        try:
            ultimo_uso = datetime.fromisoformat(ultimo_uso_txt)
            if ultimo_uso.tzinfo is None:
                ultimo_uso = ultimo_uso.replace(tzinfo=timezone.utc)
        except ValueError:
            ultimo_uso = agora
        if agora - ultimo_uso > app.permanent_session_lifetime:
            registrar_log('LOGOUT_INATIVIDADE', f'Usuário {current_user.username} saiu por inatividade')
            logout_user()
            session.clear()
            flash('Sua sessão expirou por inatividade. Entre novamente.', 'error')
            return redirect(url_for('login'))

    session.permanent = True
    session['ultimo_uso'] = agora.isoformat()
    return None


class Usuario(UserMixin):
    def __init__(self, id, username, role, nome_completo='', email='', ativo=1):
        self.id = id
        self.username = username
        self.role = role
        self.nome_completo = nome_completo
        self.email = email
        self.ativo = ativo


def get_db():
    # timeout=10 reduz falhas quando duas pessoas salvam dados quase ao mesmo tempo.
    conn = sqlite3.connect(DB_PATH, timeout=10, check_same_thread=False)
    conn.row_factory = sqlite3.Row
    conn.execute('PRAGMA foreign_keys=ON')
    # WAL melhora leitura simultanea em sistemas pequenos com varios usuarios internos.
    conn.execute('PRAGMA journal_mode=WAL')
    # NORMAL equilibra seguranca e desempenho para um SQLite usado em aplicacao web interna.
    conn.execute('PRAGMA synchronous=NORMAL')
    return conn


@contextmanager
def db_connection(commit=False):
    conn = get_db()
    try:
        yield conn
        if commit:
            conn.commit()
    except Exception:
        if commit:
            conn.rollback()
        raise
    finally:
        conn.close()


@login_manager.user_loader
def load_user(user_id):
    with db_connection() as conn:
        cols = colunas_tabela(conn, 'usuarios')
        filtro_ativo = ' AND ativo=1' if 'ativo' in cols else ''
        row = conn.execute(f'SELECT * FROM usuarios WHERE id=?{filtro_ativo}', (user_id,)).fetchone()
    if not row:
        return None
    row_keys = row.keys()
    return Usuario(row['id'], row['username'], row['role'],
                   (row['nome_completo'] if 'nome_completo' in row_keys else '') or '',
                   (row['email'] if 'email' in row_keys else '') or '',
                   row['ativo'] if 'ativo' in row_keys else 1)


def requer_papel(*papeis):
    def decorator(f):
        @wraps(f)
        def wrapped(*args, **kwargs):
            if not current_user.is_authenticated or current_user.role not in papeis:
                return jsonify({'erro': 'Acesso negado'}), 403
            return f(*args, **kwargs)
        return wrapped
    return decorator


def requer_papel_page(*papeis):
    def decorator(f):
        @wraps(f)
        def wrapped(*args, **kwargs):
            if not current_user.is_authenticated or current_user.role not in papeis:
                flash('Acesso negado.', 'error')
                return redirect(url_for('index'))
            return f(*args, **kwargs)
        return wrapped
    return decorator


# ========================================
# CONSTANTES DE NEGOCIO
# ========================================

SALAS = [
    'Consultório 1', 'Consultório 2', 'Consultório 3', 'Consultório 4',
    'Consultório 5', 'Consultório 6 (Divã)', 'Consultório 7 (Divã)',
    'Consultório 8', 'SOU / NACE', 'Ludoterapia', 'Multifuncional',
    'Sala de Grupo 1', 'Sala de Grupo 2', 'Supervisão', 'Coordenação'
]

HORARIOS = ['07:00', '08:00', '09:00', '10:00', '11:00', '12:00', '13:00',
            '14:00', '15:00', '16:00', '17:00', '18:00', '19:00', '20:00']
SALAS_RESERVAVEIS = [
    'Consultório 1', 'Consultório 2', 'Consultório 3', 'Consultório 4',
    'Consultório 5', 'Consultório 6 (Divã)', 'Consultório 7 (Divã)',
    'Consultório 8', 'Ludoterapia', 'Multifuncional',
    'Sala de Grupo 1', 'Sala de Grupo 2'
]
SALAS_COM_COMPUTADOR = [
    'Consultório 3', 'Consultório 4', 'Consultório 5',
    'Consultório 6 (Divã)', 'Consultório 7 (Divã)'
]
DIAS = ['SEGUNDA', 'TERÇA', 'QUARTA', 'QUINTA', 'SEXTA']
DIAS_PT = {
    'SEGUNDA': 'Segunda-feira', 'TERÇA': 'Terça-feira', 'QUARTA': 'Quarta-feira',
    'QUINTA': 'Quinta-feira', 'SEXTA': 'Sexta-feira',
}

CATEGORIAS = [
    'ESTAGIÁRIO 10°', 'ESTAGIÁRIO 9°',
    'SUPERVISÃO', 'NACE', 'SOU', 'MARCAR', 'NÃO MARCAR',
    'NUTRIÇÃO', 'PSICODIAGNÓSTICO', 'PSIQUIATRIA',
    'AMBULATÓRIO NEUROPSICOLOGIA', 'PLANTÃO PSICOLÓGICO',
    'PRONTUÁRIO/ESTUDAR', 'LIVRE', 'OUTRO'
]

CATEGORIAS_OPERACIONAIS = [
    c for c in CATEGORIAS
    if not c.startswith('ESTAGIÁRIO')
]

CATEGORIAS_OCUPAM_SALA = (
    'SUPERVISÃO', 'NACE', 'SOU', 'NUTRIÇÃO', 'PSICODIAGNÓSTICO',
    'PSIQUIATRIA', 'AMBULATÓRIO NEUROPSICOLOGIA', 'PLANTÃO PSICOLÓGICO',
    'PRONTUÁRIO/ESTUDAR'
)

STATUS_ATENDIMENTO = {
    '': '',
    'paciente_faltou': 'Paciente faltou',
    'profissional_desmarcou': 'Profissional desmarcou',
    'paciente_desmarcou': 'Paciente desmarcou',
}

ASSINATURA_EMAIL = (
    'Atenciosamente,\n'
    'Policlínica de Psicologia UVV\n\n'
    'Este é um e-mail automático. Não responda esta mensagem.'
)

PAPEIS_LABEL = {
    'coordenador': 'Coordenador',
    'recepcao': 'Recepcionista',
    'professor': 'Professor',
    'aluno': 'Aluno',
    'somente_leitura': 'Somente leitura'
}

LOG_RETENCAO_DIAS = 15


# ========================================
# BANCO DE DADOS
# ========================================

def normalizar_data_especifica(data_especifica):
    return agendamento_utils.normalizar_data_especifica(data_especifica, DIAS)


def dia_semana_da_data(data_especifica):
    return agendamento_utils.dia_semana_da_data(data_especifica, DIAS)


def numero_semana_sqlite(dia):
    return agendamento_utils.numero_semana_sqlite(dia, DIAS)


def validar_valores_agendamento(dia, horario, sala, categoria=''):
    return agendamento_utils.validar_valores_agendamento(
        dia, horario, sala, categoria, DIAS, HORARIOS, SALAS, CATEGORIAS
    )


def data_hoje_iso():
    return datetime.now().strftime('%Y-%m-%d')


def inteiro_query(nome, padrao, minimo=None, maximo=None):
    valor = request.args.get(nome, padrao)
    try:
        valor = int(valor)
    except (TypeError, ValueError):
        return None, f'O campo "{nome}" precisa ser um número.'
    if minimo is not None:
        valor = max(minimo, valor)
    if maximo is not None:
        valor = min(maximo, valor)
    return valor, None


def normalizar_categoria_triagem(categoria):
    return agendamento_utils.normalizar_categoria_triagem(categoria)


def texto_indica_triagem(estagiario, paciente):
    return agendamento_utils.texto_indica_triagem(estagiario, paciente)


def valor_triagem(valor, padrao=0):
    return agendamento_utils.valor_triagem(valor, padrao)


def valor_ocupa_sala(valor, padrao=None):
    return agendamento_utils.valor_ocupa_sala(valor, padrao)


def motivo_ocupacao_sala(categoria, paciente='', observacao='', data_especifica='', triagem=0):
    return agendamento_utils.motivo_ocupacao_sala(
        categoria, paciente, observacao, data_especifica, triagem, CATEGORIAS_OCUPAM_SALA
    )


def calcular_ocupa_sala(categoria, paciente='', observacao='', data_especifica='', triagem=0):
    return agendamento_utils.calcular_ocupa_sala(
        categoria, paciente, observacao, data_especifica, triagem, CATEGORIAS_OCUPAM_SALA
    )


def preparar_dados_agendamento(dados, usuario_id_padrao=None):
    return agendamento_utils.preparar_dados_agendamento(
        dados,
        usuario_id_padrao,
        DIAS,
        HORARIOS,
        SALAS,
        CATEGORIAS,
        CATEGORIAS_OCUPAM_SALA,
        STATUS_ATENDIMENTO,
    )


def usuario_pode_ver_agendamento(row):
    if not row:
        return False
    if current_user.role in ('coordenador', 'recepcao', 'somente_leitura'):
        return True
    if current_user.role == 'professor':
        return row['supervisor_id'] == current_user.id if 'supervisor_id' in row.keys() else False
    return row['usuario_id'] == current_user.id or (
        row['usuario_id'] is None and row['estagiario'] == current_user.username
    )


def agendamento_para_resposta(row):
    item = dict(row)
    if current_user.is_authenticated and current_user.role == 'aluno':
        item['paciente'] = ''
        item['observacao'] = ''
        item['paciente_label_aluno'] = 'Triagem marcada' if int(item.get('triagem') or 0) else 'Paciente marcado'
    return item


def buscar_usuario_id_aluno(username, conn):
    return usuarios_mod.buscar_usuario_id_aluno(username, conn)


def colunas_tabela(conn, tabela):
    return {r[1] for r in conn.execute(f"PRAGMA table_info({tabela})").fetchall()}


def selecionar_usuarios_para_admin(conn):
    return usuarios_mod.selecionar_usuarios_para_admin(conn, colunas_tabela)


def listar_professores_ativos(conn):
    return usuarios_mod.listar_professores_ativos(conn, colunas_tabela)


def vincular_aluno_do_agendamento(dados_ag, conn):
    aluno_id = buscar_usuario_id_aluno(dados_ag.get('estagiario'), conn)
    if aluno_id:
        dados_ag['usuario_id'] = aluno_id
    return dados_ag


def inserir_agendamento(conn, dados_ag):
    vincular_aluno_do_agendamento(dados_ag, conn)
    if 'ocupa_sala' not in dados_ag:
        dados_ag['ocupa_sala'] = calcular_ocupa_sala(
            dados_ag.get('categoria', ''),
            dados_ag.get('paciente', ''),
            dados_ag.get('observacao', ''),
            dados_ag.get('data_especifica', ''),
            dados_ag.get('triagem', 0)
        )
    cur = conn.execute(
        'INSERT INTO agendamentos(dia_semana,horario,sala,estagiario,paciente,categoria,semestre,triagem,observacao,data_especifica,usuario_id,ocupa_sala,status_atendimento)'
        ' VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?)',
        (
            dados_ag['dia'], dados_ag['horario'], dados_ag['sala'],
            dados_ag['estagiario'], dados_ag['paciente'], dados_ag['categoria'],
            dados_ag['semestre'], dados_ag['triagem'], dados_ag['observacao'],
            dados_ag['data_especifica'], dados_ag['usuario_id'], dados_ag['ocupa_sala'],
            dados_ag.get('status_atendimento', '')
        )
    )
    return cur.lastrowid


def config_email():
    return {
        'host': SMTP_HOST,
        'port': SMTP_PORT,
        'user': SMTP_USER,
        'password': SMTP_PASSWORD,
        'tls': SMTP_TLS,
        'from': EMAIL_FROM,
        'base_url': EMAIL_BASE_URL,
    }


def email_configurado():
    return email_utils.email_configurado(config_email())


def diagnostico_smtp():
    return email_utils.diagnostico_smtp(config_email())


def url_absoluta(endpoint, **valores):
    if EMAIL_BASE_URL:
        return f"{EMAIL_BASE_URL}{url_for(endpoint, **valores)}"
    return url_for(endpoint, _external=True, **valores)


def email_de_teste_ou_invalido(email):
    return email_utils.email_de_teste_ou_invalido(email)


def validar_email_usuario(email, obrigatorio=False):
    return email_utils.validar_email_usuario(email, obrigatorio)


def enviar_email(destinatario, assunto, corpo):
    destinatario = (destinatario or '').strip()
    if not destinatario:
        return False
    if email_de_teste_ou_invalido(destinatario):
        registrar_log('EMAIL_IGNORADO', f'E-mail de teste/inválido ignorado: {destinatario} | {assunto}')
        return False
    if not email_configurado():
        registrar_log('EMAIL_NAO_CONFIGURADO', f'E-mail não enviado para {destinatario}: {assunto}')
        return False

    try:
        email_utils.enviar_email_smtp(config_email(), destinatario, assunto, corpo)
        return True
    except Exception as exc:
        registrar_log('EMAIL_ERRO', f'Falha ao enviar para {destinatario}: {assunto} | {exc}')
        return False


def enviar_email_multiplos(destinatarios, assunto, corpo):
    enviados = 0
    for destinatario in sorted(set(d for d in destinatarios if d)):
        if enviar_email(destinatario, assunto, corpo):
            enviados += 1
    return enviados


def emails_usuarios_por_papel(*papeis):
    if not papeis:
        return []
    placeholders = ','.join('?' for _ in papeis)
    conn = get_db()
    try:
        rows = conn.execute(
            f"""
            SELECT email
            FROM usuarios
            WHERE ativo=1
              AND TRIM(COALESCE(email, ''))!=''
              AND role IN ({placeholders})
            """,
            papeis
        ).fetchall()
    finally:
        conn.close()
    return [r['email'] for r in rows]


def nome_exibicao_usuario(usuario_row):
    if not usuario_row:
        return ''
    nome = usuario_row['nome_completo'] if 'nome_completo' in usuario_row.keys() else ''
    return nome or usuario_row['username']


def corpo_email_padrao(saudacao, mensagem, titulo_detalhes='', detalhes=None, observacao='', encerramento=''):
    partes = [saudacao.strip(), mensagem.strip()]
    detalhes = detalhes or []
    if titulo_detalhes and detalhes:
        partes.append(titulo_detalhes.strip())
        partes.append('\n'.join(f'• {rotulo}: {valor or "-"}' for rotulo, valor in detalhes))
    if observacao:
        partes.append(observacao.strip())
    if encerramento:
        partes.append(encerramento.strip())
    partes.append(ASSINATURA_EMAIL)
    return '\n\n'.join(p for p in partes if p)


def formatar_data_email(dados_ag):
    data_ref = (dados_ag.get('data_especifica') or '').strip()
    if data_ref:
        try:
            data_obj = datetime.strptime(data_ref, '%Y-%m-%d')
            dia_nome = DIAS_PT.get(DIAS[data_obj.weekday()], '')
            return f'{data_obj.strftime("%d/%m/%Y")} ({dia_nome})' if dia_nome else data_obj.strftime('%d/%m/%Y')
        except (ValueError, TypeError):
            return data_ref

    dia = dados_ag.get('dia') or dados_ag.get('dia_semana') or ''
    return DIAS_PT.get(dia, dia or '-')


def hash_token(token):
    return hashlib.sha256(token.encode('utf-8')).hexdigest()


def criar_token_email(conn, usuario_id, tipo, horas=24):
    token = secrets.token_urlsafe(32)
    expira_em = (datetime.now(timezone.utc) + timedelta(hours=horas)).strftime('%Y-%m-%d %H:%M:%S')
    conn.execute(
        """
        INSERT INTO tokens_email(usuario_id, tipo, token_hash, expira_em)
        VALUES(?,?,?,?)
        """,
        (usuario_id, tipo, hash_token(token), expira_em)
    )
    return token


def buscar_token_email(token, tipo):
    conn = get_db()
    try:
        row = conn.execute(
            """
            SELECT t.*, u.username, u.email, u.nome_completo
            FROM tokens_email t
            JOIN usuarios u ON u.id = t.usuario_id
            WHERE t.token_hash=? AND t.tipo=? AND t.usado_em IS NULL AND t.expira_em > datetime('now')
            """,
            (hash_token(token), tipo)
        ).fetchone()
    finally:
        conn.close()
    return row


def marcar_token_usado(conn, token_id):
    conn.execute('UPDATE tokens_email SET usado_em=CURRENT_TIMESTAMP WHERE id=?', (token_id,))


def preparar_convite_criacao_conta(conn, usuario_row):
    if not usuario_row or not usuario_row['email']:
        return None
    token = criar_token_email(conn, usuario_row['id'], 'convite', horas=72)
    link = url_absoluta('redefinir_senha', token=token)
    return (
        usuario_row['email'],
        'Convite para acessar o Mapa de Sala',
        corpo_email_padrao(
            f'Olá, {nome_exibicao_usuario(usuario_row)}!',
            'Uma conta foi criada para você no Mapa de Sala da Policlínica de Psicologia UVV.',
            observacao=(
                'Para acessar o sistema pela primeira vez, crie sua senha utilizando o link abaixo:\n\n'
                f'{link}\n\n'
                'Este link é válido por 72 horas.\n'
                'Após definir sua senha, você poderá acessar o sistema normalmente.\n'
                'Caso não reconheça esta solicitação, ignore este e-mail.'
            )
        )
    )


def preparar_aviso_conta_criada(usuario_row):
    if not usuario_row or not usuario_row['email']:
        return None
    link = url_absoluta('login')
    return (
        usuario_row['email'],
        'Conta criada no Mapa de Sala',
        corpo_email_padrao(
            f'Olá, {nome_exibicao_usuario(usuario_row)}!',
            'Sua conta no Mapa de Sala foi criada pela coordenação.',
            observacao=(
                f'Acesse o sistema por este link:\n{link}\n\n'
                'Se você ainda não recebeu sua senha, procure a coordenação ou use a opção de recuperar senha.'
            )
        )
    )


def notificar_senha_alterada_email(usuario_row, origem):
    if not usuario_row or not usuario_row['email']:
        return False
    return enviar_email(
        usuario_row['email'],
        'Senha alterada com sucesso',
        corpo_email_padrao(
            f'Olá, {nome_exibicao_usuario(usuario_row)}!',
            'Sua senha de acesso ao Mapa de Sala foi alterada com sucesso.',
            observacao='Caso você não reconheça essa alteração, entre em contato imediatamente com o administrador do sistema.'
        )
    )


def resumo_agendamento_email(dados_ag):
    tipo = 'Triagem' if int(dados_ag.get('triagem') or 0) else 'Atendimento'
    status = STATUS_ATENDIMENTO.get(dados_ag.get('status_atendimento') or '')
    if status:
        situacao = status
    else:
        situacao = 'Com paciente vinculado' if (dados_ag.get('paciente') or '').strip() else 'Sem paciente vinculado'
    return [
        ('Tipo de agendamento', tipo),
        ('Situação', situacao),
        ('Data', formatar_data_email(dados_ag)),
        ('Horário', dados_ag.get('horario')),
        ('Sala', dados_ag.get('sala')),
        ('Categoria', dados_ag.get('categoria') or 'Sem categoria'),
    ]


def classificar_evento_agendamento(dados_ag, acao, dados_antes=None):
    if acao == 'excluido':
        return 'cancelado'
    if dados_ag.get('status_atendimento'):
        return 'cancelado'
    if acao == 'alterado' and dados_antes:
        paciente_antes = (dados_antes.get('paciente') or '').strip()
        paciente_depois = (dados_ag.get('paciente') or '').strip()
        if not paciente_antes and paciente_depois:
            return 'paciente_vinculado'
        if paciente_antes and not paciente_depois:
            return 'paciente_removido'
    return acao


def conteudo_agendamento_email(evento, dados_ag):
    detalhes = resumo_agendamento_email(dados_ag)
    privacidade = 'Por questões de privacidade, este e-mail não exibe o nome do paciente.'
    mapa = {
        'criado': {
            'assunto': 'Novo agendamento cadastrado',
            'mensagem': 'Um novo agendamento foi registrado para você no Mapa de Sala.',
            'titulo': 'Detalhes do agendamento',
            'observacao': f'{privacidade}\nPara visualizar mais informações, acesse o sistema.',
        },
        'alterado': {
            'assunto': 'Agendamento atualizado',
            'mensagem': 'Um agendamento vinculado à sua agenda foi atualizado.',
            'titulo': 'Novas informações',
            'observacao': f'{privacidade}\nPara visualizar os detalhes da alteração, acesse o sistema.',
        },
        'cancelado': {
            'assunto': 'Agendamento cancelado',
            'mensagem': 'Um agendamento foi cancelado no Mapa de Sala.',
            'titulo': 'Informações do agendamento',
            'observacao': 'Caso tenha dúvidas, consulte o sistema.',
        },
        'paciente_vinculado': {
            'assunto': 'Paciente vinculado ao seu horário',
            'mensagem': 'Um paciente foi vinculado a um horário disponível em sua agenda.',
            'titulo': 'Informações',
            'observacao': f'{privacidade}\nAcesse o sistema para visualizar as informações completas.',
        },
        'paciente_removido': {
            'assunto': 'Paciente desvinculado do agendamento',
            'mensagem': 'O paciente anteriormente vinculado a um de seus horários foi removido.\nO horário voltou a ficar disponível.',
            'titulo': 'Informações',
            'observacao': 'Caso tenha dúvidas, consulte o sistema.',
        },
    }
    conteudo = dict(mapa.get(evento, mapa['alterado']))
    conteudo['detalhes'] = detalhes
    return conteudo


def notificar_agendamento_email(dados_ag, acao, dados_antes=None):
    aluno_id = dados_ag.get('usuario_id')
    aluno_username = dados_ag.get('estagiario') or ''
    conn = get_db()
    try:
        aluno = None
        if aluno_id:
            aluno = conn.execute('SELECT * FROM usuarios WHERE id=?', (aluno_id,)).fetchone()
        if not aluno and aluno_username:
            aluno = conn.execute(
                "SELECT * FROM usuarios WHERE username=? AND role='aluno'",
                (aluno_username,)
            ).fetchone()

        professor = None
        if aluno and aluno['supervisor_id']:
            professor = conn.execute('SELECT * FROM usuarios WHERE id=?', (aluno['supervisor_id'],)).fetchone()
    finally:
        conn.close()

    evento = classificar_evento_agendamento(dados_ag, acao, dados_antes)
    conteudo = conteudo_agendamento_email(evento, dados_ag)
    corpo_base = corpo_email_padrao(
        'Olá!',
        conteudo['mensagem'],
        conteudo['titulo'],
        conteudo['detalhes'],
        conteudo['observacao']
    )

    if aluno and aluno['email']:
        enviar_email(aluno['email'], conteudo['assunto'], corpo_base)
    if professor and professor['email']:
        enviar_email(
            professor['email'],
            f'{conteudo["assunto"]} para aluno supervisionado',
            f'Aluno: {aluno["username"]}\n\n{corpo_base}'
        )


def notificar_reserva_solicitada_email(reserva):
    if not reserva:
        return 0
    supervisor_nome = ''
    if reserva.get('usuario_id'):
        conn = get_db()
        try:
            supervisor = conn.execute(
                """
                SELECT COALESCE(NULLIF(p.nome_completo, ''), p.username, '') AS nome
                FROM usuarios u
                LEFT JOIN usuarios p ON p.id = u.supervisor_id
                WHERE u.id=?
                """,
                (reserva.get('usuario_id'),)
            ).fetchone()
            supervisor_nome = supervisor['nome'] if supervisor else ''
        finally:
            conn.close()
    tipo = 'sala' if reserva.get('tipo') == 'sala' else 'teste/instrumento'
    detalhe = reserva.get('tipo_sala') if reserva.get('tipo') == 'sala' else reserva.get('instrumento')
    corpo = corpo_email_padrao(
        'Olá!',
        f'Uma nova reserva de {tipo} foi solicitada no Mapa de Sala.',
        'Informações da reserva',
        [
            ('Aluno', reserva.get('usuario')),
            ('Supervisor', supervisor_nome or '-'),
            ('Data', formatar_data_email({'data_especifica': reserva.get('data_uso')})),
            ('Horário', f'{reserva.get("horario_inicio")}{(" até " + reserva.get("horario_fim")) if reserva.get("horario_fim") else ""}'),
            ('Detalhe', detalhe or '-'),
            ('Finalidade', reserva.get('finalidade') or '-'),
        ],
        'Acesse a tela de Reservas para aprovar ou recusar.'
    )
    return enviar_email_multiplos(
        emails_usuarios_por_papel('coordenador', 'recepcao'),
        'Nova reserva aguardando análise - Mapa de Sala',
        corpo
    )


def notificar_reserva_email(reserva, status, resposta=''):
    if not reserva or not reserva.get('usuario_id'):
        return False
    conn = get_db()
    try:
        usuario = conn.execute('SELECT * FROM usuarios WHERE id=?', (reserva['usuario_id'],)).fetchone()
    finally:
        conn.close()
    if not usuario or not usuario['email']:
        return False

    tipo = 'sala' if reserva.get('tipo') == 'sala' else 'teste/instrumento'
    status_txt = {
        'aprovada': 'aprovada',
        'recusada': 'recusada',
        'separado': 'com instrumento separado',
        'guardado': 'com instrumento guardado/devolvido',
    }.get(status, status)
    detalhe = reserva.get('sala_atribuida') if reserva.get('tipo') == 'sala' else reserva.get('instrumento')
    corpo = corpo_email_padrao(
        f'Olá, {nome_exibicao_usuario(usuario)}!',
        f'Sua reserva de {tipo} foi {status_txt}.',
        'Informações da reserva',
        [
            ('Data', formatar_data_email({'data_especifica': reserva.get('data_uso')})),
            ('Horário', f'{reserva.get("horario_inicio")}{(" até " + reserva.get("horario_fim")) if reserva.get("horario_fim") else ""}'),
            ('Detalhe', detalhe or '-'),
        ],
        f'Mensagem da recepção/coordenação: {resposta}' if resposta else 'Acesse o sistema para acompanhar a solicitação.'
    )
    return enviar_email(usuario['email'], f'Reserva {status_txt} - Mapa de Sala', corpo)


def formatar_data_iso(data_iso):
    try:
        return datetime.strptime(data_iso or '', '%Y-%m-%d').strftime('%d/%m/%Y')
    except (TypeError, ValueError):
        return data_iso or '-'


def montar_detalhes_coord_agendamento(agendamento):
    horario = agendamento.get('horario_inicio') or ''
    if agendamento.get('horario_fim'):
        horario = f'{horario} até {agendamento.get("horario_fim")}'
    return [
        ('Aluno', agendamento.get('aluno_nome') or '-'),
        ('Data', formatar_data_iso(agendamento.get('data_disponivel'))),
        ('Horário', horario or '-'),
        ('Local', agendamento.get('local') or '-'),
        ('Assunto', agendamento.get('assunto') or '-'),
        ('Observação', agendamento.get('observacao') or '-'),
        ('Status', agendamento.get('status') or '-'),
    ]


def notificar_coord_agendamento_email(agendamento, evento, resposta=''):
    if not agendamento:
        return 0

    assunto_aluno = {
        'criado': 'Horário com a coordenação confirmado',
        'cancelado': 'Horário com a coordenação cancelado',
        'alterado': 'Horário com a coordenação atualizado',
    }.get(evento, 'Horário com a coordenação atualizado')
    mensagem_aluno = {
        'criado': 'Seu horário para conversar com a coordenação foi confirmado.',
        'cancelado': 'Seu horário para conversar com a coordenação foi cancelado.',
        'alterado': 'Seu horário para conversar com a coordenação foi atualizado.',
    }.get(evento, 'Seu horário com a coordenação foi atualizado.')

    enviados = 0
    aluno_email = agendamento.get('aluno_email') or ''
    if aluno_email:
        corpo_aluno = corpo_email_padrao(
            f'Olá, {agendamento.get("aluno_nome") or "aluno"}!',
            mensagem_aluno,
            'Informações do horário',
            montar_detalhes_coord_agendamento(agendamento),
            f'Mensagem da coordenação: {resposta}' if resposta else 'Acesse o sistema para acompanhar seus horários.'
        )
        if enviar_email(aluno_email, assunto_aluno, corpo_aluno):
            enviados += 1

    if evento == 'criado':
        corpo_coord = corpo_email_padrao(
            'Olá!',
            'Um aluno reservou um horário para conversar com a coordenação.',
            'Informações do horário',
            montar_detalhes_coord_agendamento(agendamento),
            'Acesse a tela Coordenação para acompanhar.'
        )
        enviados += enviar_email_multiplos(
            emails_usuarios_por_papel('coordenador'),
            'Novo horário reservado com a coordenação',
            corpo_coord
        )

    return enviados


def limpar_logs_antigos(conn):
    """Remove historico antigo para impedir crescimento continuo do arquivo SQLite."""
    cur = conn.execute(
        "DELETE FROM historico WHERE ts < datetime('now', '-' || ? || ' days')",
        (LOG_RETENCAO_DIAS,)
    )
    return cur.rowcount


def criar_indices_agendamentos(conn):
    conn.executescript(
        "DROP INDEX IF EXISTS idx_conflito;"
        "CREATE UNIQUE INDEX IF NOT EXISTS idx_conflito_semanal "
        "ON agendamentos(dia_semana, horario, sala) "
        "WHERE ocupa_sala = 1 AND (data_especifica IS NULL OR data_especifica = '');"
        "CREATE UNIQUE INDEX IF NOT EXISTS idx_conflito_data "
        "ON agendamentos(data_especifica, horario, sala) "
        "WHERE ocupa_sala = 1 AND data_especifica IS NOT NULL AND data_especifica != '';"
        "CREATE INDEX IF NOT EXISTS idx_dia_semana ON agendamentos(dia_semana);"
        "CREATE INDEX IF NOT EXISTS idx_data_especifica ON agendamentos(data_especifica);"
        "CREATE INDEX IF NOT EXISTS idx_ocupa_sala ON agendamentos(ocupa_sala);"
    )


def migrar_fk_usuario_id_agendamentos(conn):
    fks = conn.execute("PRAGMA foreign_key_list(agendamentos)").fetchall()
    if any(row['table'] == 'usuarios' and row['from'] == 'usuario_id' for row in fks):
        return

    cols = [r[1] for r in conn.execute("PRAGMA table_info(agendamentos)").fetchall()]
    if 'usuario_id' not in cols:
        return

    conn.executescript(
        "DROP INDEX IF EXISTS idx_conflito;"
        "DROP INDEX IF EXISTS idx_conflito_semanal;"
        "DROP INDEX IF EXISTS idx_conflito_data;"
        "DROP INDEX IF EXISTS idx_dia_semana;"
        "DROP INDEX IF EXISTS idx_data_especifica;"
        "DROP INDEX IF EXISTS idx_ocupa_sala;"
        "ALTER TABLE agendamentos RENAME TO agendamentos_old;"
        "CREATE TABLE agendamentos ("
        "id INTEGER PRIMARY KEY AUTOINCREMENT,"
        "dia_semana TEXT DEFAULT 'SEGUNDA',"
        "horario TEXT NOT NULL,"
        "sala TEXT NOT NULL,"
        "estagiario TEXT DEFAULT '',"
        "paciente TEXT DEFAULT '',"
        "categoria TEXT DEFAULT '',"
        "semestre INTEGER DEFAULT 0,"
        "triagem INTEGER DEFAULT 0,"
        "observacao TEXT DEFAULT '',"
        "data_especifica TEXT DEFAULT '',"
        "usuario_id INTEGER DEFAULT NULL REFERENCES usuarios(id) ON DELETE SET NULL,"
        "ocupa_sala INTEGER DEFAULT 0,"
        "status_atendimento TEXT DEFAULT '',"
        "created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,"
        "updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP"
        ");"
        "INSERT INTO agendamentos("
        "id,dia_semana,horario,sala,estagiario,paciente,categoria,semestre,triagem,"
        "observacao,data_especifica,usuario_id,ocupa_sala,status_atendimento,created_at,updated_at"
        ") SELECT "
        "id,dia_semana,horario,sala,estagiario,paciente,categoria,semestre,triagem,"
        "observacao,data_especifica,"
        "CASE WHEN usuario_id IS NULL OR EXISTS (SELECT 1 FROM usuarios u WHERE u.id = agendamentos_old.usuario_id) "
        "THEN usuario_id ELSE NULL END,"
        "ocupa_sala,"
        "COALESCE(status_atendimento, ''),"
        "created_at,updated_at "
        "FROM agendamentos_old;"
        "DROP TABLE agendamentos_old;"
    )


def corrigir_vinculos_alunos_agendamentos(conn):
    conn.execute(
        """
        UPDATE agendamentos
        SET usuario_id = (
            SELECT u.id
            FROM usuarios u
            WHERE u.username = agendamentos.estagiario
              AND u.role = 'aluno'
              AND u.ativo = 1
        )
        WHERE EXISTS (
            SELECT 1
            FROM usuarios u
            WHERE u.username = agendamentos.estagiario
              AND u.role = 'aluno'
              AND u.ativo = 1
        )
        """
    )


def migrar_categorias_triagem(conn):
    conn.execute(
        """
        UPDATE agendamentos
        SET categoria = 'ESTAGIÁRIO 10°',
            triagem = 1
        WHERE categoria = 'ESTAGIÁRIO 10° TRIAGEM'
        """
    )
    conn.execute(
        """
        UPDATE agendamentos
        SET categoria = 'ESTAGIÁRIO 9°',
            triagem = 1
        WHERE categoria = 'ESTAGIÁRIO 9° TRIAGEM'
        """
    )


def recalcular_ocupacao_sala_agendamentos(conn):
    placeholders = ','.join('?' for _ in CATEGORIAS_OCUPAM_SALA)
    conn.execute(
        f"""
        UPDATE agendamentos
        SET ocupa_sala = CASE
            WHEN TRIM(COALESCE(status_atendimento, '')) != '' THEN 0
            WHEN TRIM(COALESCE(paciente, '')) != '' THEN 1
            WHEN categoria IN ({placeholders}) THEN 1
            WHEN TRIM(COALESCE(data_especifica, '')) != ''
                 AND TRIM(COALESCE(observacao, '')) != '' THEN 1
            ELSE 0
        END
        """,
        CATEGORIAS_OCUPAM_SALA
    )


def criar_tabelas_base(conn):
    conn.executescript(
        "CREATE TABLE IF NOT EXISTS agendamentos ("
        "id INTEGER PRIMARY KEY AUTOINCREMENT,"
        "dia_semana TEXT DEFAULT 'SEGUNDA',"
        "horario TEXT NOT NULL,"
        "sala TEXT NOT NULL,"
        "estagiario TEXT DEFAULT '',"
        "paciente TEXT DEFAULT '',"
        "categoria TEXT DEFAULT '',"
        "semestre INTEGER DEFAULT 0,"
        "triagem INTEGER DEFAULT 0,"
        "observacao TEXT DEFAULT '',"
        "data_especifica TEXT DEFAULT '',"
        "usuario_id INTEGER DEFAULT NULL REFERENCES usuarios(id) ON DELETE SET NULL,"
        "ocupa_sala INTEGER DEFAULT 0,"
        "status_atendimento TEXT DEFAULT '',"
        "created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,"
        "updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP"
        ");"
        "CREATE TABLE IF NOT EXISTS historico ("
        "id INTEGER PRIMARY KEY AUTOINCREMENT,"
        "usuario TEXT DEFAULT '',"
        "acao TEXT,"
        "dados TEXT,"
        "ip TEXT DEFAULT '',"
        "user_agent TEXT DEFAULT '',"
        "ts TIMESTAMP DEFAULT CURRENT_TIMESTAMP"
        ");"
        "CREATE TABLE IF NOT EXISTS usuarios ("
        "id INTEGER PRIMARY KEY AUTOINCREMENT,"
        "username TEXT NOT NULL UNIQUE,"
        "password_hash TEXT NOT NULL,"
        "role TEXT NOT NULL DEFAULT 'aluno',"
        "nome_completo TEXT DEFAULT '',"
        "email TEXT DEFAULT '',"
        "supervisor_id INTEGER DEFAULT NULL REFERENCES usuarios(id) ON DELETE SET NULL,"
        "ativo INTEGER DEFAULT 1,"
        "created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP"
        ");"
        "CREATE TABLE IF NOT EXISTS tokens_email ("
        "id INTEGER PRIMARY KEY AUTOINCREMENT,"
        "usuario_id INTEGER NOT NULL REFERENCES usuarios(id) ON DELETE CASCADE,"
        "tipo TEXT NOT NULL,"
        "token_hash TEXT NOT NULL UNIQUE,"
        "expira_em TIMESTAMP NOT NULL,"
        "usado_em TIMESTAMP DEFAULT NULL,"
        "created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP"
        ");"
        "CREATE TABLE IF NOT EXISTS reservas ("
        "id INTEGER PRIMARY KEY AUTOINCREMENT,"
        "usuario_id INTEGER DEFAULT NULL REFERENCES usuarios(id) ON DELETE SET NULL,"
        "usuario TEXT DEFAULT '',"
        "tipo TEXT NOT NULL,"
        "status TEXT DEFAULT 'pendente',"
        "data_uso TEXT NOT NULL,"
        "horario_inicio TEXT DEFAULT '',"
        "horario_fim TEXT DEFAULT '',"
        "tipo_sala TEXT DEFAULT '',"
        "sala_atribuida TEXT DEFAULT '',"
        "instrumento TEXT DEFAULT '',"
        "finalidade TEXT DEFAULT '',"
        "observacao TEXT DEFAULT '',"
        "resposta TEXT DEFAULT '',"
        "agendamento_ids TEXT DEFAULT '',"
        "analisado_por TEXT DEFAULT '',"
        "created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,"
        "updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP"
        ");"
        "CREATE INDEX IF NOT EXISTS idx_reservas_status ON reservas(status);"
        "CREATE INDEX IF NOT EXISTS idx_reservas_usuario ON reservas(usuario_id);"
        "CREATE TABLE IF NOT EXISTS tarefas_painel ("
        "id INTEGER PRIMARY KEY AUTOINCREMENT,"
        "titulo TEXT NOT NULL,"
        "detalhe TEXT DEFAULT '',"
        "criado_por TEXT DEFAULT '',"
        "created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP"
        ");"
        "CREATE TABLE IF NOT EXISTS solicitacoes_vagas ("
        "id INTEGER PRIMARY KEY AUTOINCREMENT,"
        "professor_id INTEGER DEFAULT NULL REFERENCES usuarios(id) ON DELETE SET NULL,"
        "professor_nome TEXT DEFAULT '',"
        "aluno_id INTEGER DEFAULT NULL REFERENCES usuarios(id) ON DELETE SET NULL,"
        "aluno_nome TEXT DEFAULT '',"
        "vagas_paciente INTEGER DEFAULT 0,"
        "vagas_triagem INTEGER DEFAULT 0,"
        "vagas_paciente_liberadas INTEGER DEFAULT 0,"
        "vagas_triagem_liberadas INTEGER DEFAULT 0,"
        "observacao TEXT DEFAULT '',"
        "status TEXT DEFAULT 'pendente',"
        "resposta TEXT DEFAULT '',"
        "analisado_por TEXT DEFAULT '',"
        "created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,"
        "updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP"
        ");"
        "CREATE INDEX IF NOT EXISTS idx_solicitacoes_vagas_status ON solicitacoes_vagas(status);"
        "CREATE INDEX IF NOT EXISTS idx_solicitacoes_vagas_professor ON solicitacoes_vagas(professor_id);"
        "CREATE TABLE IF NOT EXISTS coordenacao_horarios ("
        "id INTEGER PRIMARY KEY AUTOINCREMENT,"
        "coordenador_id INTEGER DEFAULT NULL REFERENCES usuarios(id) ON DELETE SET NULL,"
        "coordenador_nome TEXT DEFAULT '',"
        "data_disponivel TEXT NOT NULL,"
        "horario_inicio TEXT NOT NULL,"
        "horario_fim TEXT DEFAULT '',"
        "local TEXT DEFAULT '',"
        "observacao TEXT DEFAULT '',"
        "ativo INTEGER DEFAULT 1,"
        "created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,"
        "updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP"
        ");"
        "CREATE TABLE IF NOT EXISTS coordenacao_agendamentos ("
        "id INTEGER PRIMARY KEY AUTOINCREMENT,"
        "horario_id INTEGER NOT NULL REFERENCES coordenacao_horarios(id) ON DELETE CASCADE,"
        "aluno_id INTEGER DEFAULT NULL REFERENCES usuarios(id) ON DELETE SET NULL,"
        "aluno_nome TEXT DEFAULT '',"
        "assunto TEXT DEFAULT '',"
        "observacao TEXT DEFAULT '',"
        "status TEXT DEFAULT 'confirmado',"
        "resposta TEXT DEFAULT '',"
        "created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,"
        "updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP"
        ");"
        "CREATE INDEX IF NOT EXISTS idx_coord_horarios_data ON coordenacao_horarios(data_disponivel, horario_inicio);"
        "CREATE INDEX IF NOT EXISTS idx_coord_ag_horario ON coordenacao_agendamentos(horario_id);"
        "CREATE INDEX IF NOT EXISTS idx_coord_ag_aluno ON coordenacao_agendamentos(aluno_id);"
        "CREATE TABLE IF NOT EXISTS backups_importacao ("
        "id INTEGER PRIMARY KEY AUTOINCREMENT,"
        "usuario TEXT DEFAULT '',"
        "tipo TEXT DEFAULT '',"
        "arquivo TEXT DEFAULT '',"
        "total_agendamentos INTEGER DEFAULT 0,"
        "dados_json TEXT NOT NULL,"
        "created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP"
        ");"
        "CREATE TABLE IF NOT EXISTS schema_migrations ("
        "version TEXT PRIMARY KEY,"
        "description TEXT DEFAULT '',"
        "applied_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP"
        ");"
    )


def criar_usuario_coordenador_padrao(conn):
    existe = conn.execute("SELECT id FROM usuarios WHERE username='coordenador'").fetchone()
    if existe:
        return
    conn.execute(
        "INSERT INTO usuarios(username, password_hash, role) VALUES(?,?,?)",
        ('coordenador', generate_password_hash('mudar@2026'), 'coordenador')
    )


SUPERVISORES_PADRAO = (
    ('fabio.pereira', 'Fabio Nogueira Pereira'),
    ('roger.machado', 'Roger Elias Bernabé Machado'),
    ('eduardo.lopes', 'Eduardo Barbosa Lopes'),
    ('rodrigo.salgado', 'Rodrigo Cruvinel Salgado'),
    ('luanza.mai', 'Luanza Pavesi Mai'),
    ('cleilson.reis', 'Cleilson Teobaldo Reis'),
    ('simone.pylro', 'Simone Chabudee Pylro'),
    ('hildiceia.affonso', 'Hildiceia dos Santos Affonso'),
    ('christiane.ronchete', 'Christiane Furlan Ronchete'),
    ('jaqueline.bagalho', 'Jaqueline Oliveira Bagalho'),
)


def criar_supervisores_padrao(conn):
    for username, nome_completo in SUPERVISORES_PADRAO:
        existente_por_nome = conn.execute(
            """
            SELECT id
            FROM usuarios
            WHERE role='professor'
              AND LOWER(TRIM(nome_completo))=LOWER(TRIM(?))
            """,
            (nome_completo,)
        ).fetchone()
        if existente_por_nome:
            username_em_uso = conn.execute(
                'SELECT id FROM usuarios WHERE username=? AND id!=?',
                (username, existente_por_nome['id'])
            ).fetchone()
            if not username_em_uso:
                conn.execute(
                    "UPDATE usuarios SET username=?, nome_completo=?, ativo=1 WHERE id=?",
                    (username, nome_completo, existente_por_nome['id'])
                )
                continue
            conn.execute(
                "UPDATE usuarios SET nome_completo=?, ativo=1 WHERE id=?",
                (nome_completo, existente_por_nome['id'])
            )
            continue

        username_final = username
        sufixo = 2
        while conn.execute('SELECT id FROM usuarios WHERE username=?', (username_final,)).fetchone():
            username_final = f'{username}.{sufixo}'
            sufixo += 1

        senha_aleatoria = secrets.token_urlsafe(24)
        conn.execute(
            """
            INSERT INTO usuarios(username, nome_completo, password_hash, role, ativo)
            VALUES(?,?,?,?,1)
            """,
            (username_final, nome_completo, generate_password_hash(senha_aleatoria), 'professor')
        )


def aplicar_migracoes_simples(conn):
    ocupa_col_criada = adicionar_coluna_se_ausente(
        conn,
        'agendamentos',
        'ocupa_sala',
        "ALTER TABLE agendamentos ADD COLUMN ocupa_sala INTEGER DEFAULT 0"
    )
    adicionar_coluna_se_ausente(
        conn,
        'agendamentos',
        'usuario_id',
        "ALTER TABLE agendamentos ADD COLUMN usuario_id INTEGER DEFAULT NULL"
    )
    adicionar_coluna_se_ausente(
        conn,
        'agendamentos',
        'status_atendimento',
        "ALTER TABLE agendamentos ADD COLUMN status_atendimento TEXT DEFAULT ''"
    )
    adicionar_coluna_se_ausente(
        conn,
        'usuarios',
        'nome_completo',
        "ALTER TABLE usuarios ADD COLUMN nome_completo TEXT DEFAULT ''"
    )
    adicionar_coluna_se_ausente(conn, 'usuarios', 'email', "ALTER TABLE usuarios ADD COLUMN email TEXT DEFAULT ''")
    adicionar_coluna_se_ausente(conn, 'usuarios', 'ativo', "ALTER TABLE usuarios ADD COLUMN ativo INTEGER DEFAULT 1")
    adicionar_coluna_se_ausente(
        conn,
        'usuarios',
        'supervisor_id',
        "ALTER TABLE usuarios ADD COLUMN supervisor_id INTEGER DEFAULT NULL"
    )
    adicionar_coluna_se_ausente(
        conn,
        'usuarios',
        'created_at',
        "ALTER TABLE usuarios ADD COLUMN created_at TEXT DEFAULT ''"
    )
    adicionar_coluna_se_ausente(conn, 'historico', 'ip', "ALTER TABLE historico ADD COLUMN ip TEXT DEFAULT ''")
    adicionar_coluna_se_ausente(
        conn,
        'historico',
        'user_agent',
        "ALTER TABLE historico ADD COLUMN user_agent TEXT DEFAULT ''"
    )
    adicionar_coluna_se_ausente(
        conn,
        'solicitacoes_vagas',
        'vagas_paciente_liberadas',
        "ALTER TABLE solicitacoes_vagas ADD COLUMN vagas_paciente_liberadas INTEGER DEFAULT 0"
    )
    adicionar_coluna_se_ausente(
        conn,
        'solicitacoes_vagas',
        'vagas_triagem_liberadas',
        "ALTER TABLE solicitacoes_vagas ADD COLUMN vagas_triagem_liberadas INTEGER DEFAULT 0"
    )
    return ocupa_col_criada


def aplicar_migracoes_dados(conn):
    migrar_fk_usuario_id_agendamentos(conn)
    corrigir_vinculos_alunos_agendamentos(conn)
    executar_migration(
        conn,
        '2026-06-30-categorias-triagem',
        'Normaliza categorias antigas de triagem',
        migrar_categorias_triagem
    )
    executar_migration(
        conn,
        '2026-07-01-recalcular-ocupacao-sala',
        'Recalcula ocupa_sala com regra centralizada',
        recalcular_ocupacao_sala_agendamentos
    )
    executar_migration(
        conn,
        '2026-07-02-supervisores-padrao',
        'Cadastra professores supervisores padrao',
        criar_supervisores_padrao
    )
    executar_migration(
        conn,
        '2026-07-02-supervisores-nomes-completos',
        'Atualiza usernames dos supervisores para nomes completos',
        criar_supervisores_padrao
    )
    executar_migration(
        conn,
        '2026-07-02-supervisores-usuarios-curtos',
        'Atualiza usernames dos supervisores para primeiro.ultimo',
        criar_supervisores_padrao
    )


def init_db():
    with db_connection(commit=True) as conn:
        criar_tabelas_base(conn)
        criar_usuario_coordenador_padrao(conn)
        aplicar_migracoes_simples(conn)
        aplicar_migracoes_dados(conn)
        criar_indices_agendamentos(conn)
        limpar_logs_antigos(conn)


def registrar_log(acao, dados=''):
    if has_request_context():
        usuario = current_user.username if current_user.is_authenticated else 'sistema'
        ip = request.headers.get('X-Forwarded-For', request.remote_addr or '').split(',')[0].strip()
        user_agent = (request.headers.get('User-Agent') or '')[:300]
    else:
        usuario = 'sistema'
        ip = 'console'
        user_agent = 'flask-cli'
    with db_connection(commit=True) as conn:
        conn.execute(
            'INSERT INTO historico(usuario, acao, dados, ip, user_agent) VALUES(?,?,?,?,?)',
            (usuario, acao, dados, ip, user_agent)
        )
        limpar_logs_antigos(conn)


def coletar_saude_sistema():
    with db_connection() as conn:
        return system_utils.coletar_saude_sistema(conn, VERSAO, diagnostico_smtp())


def descricao_agendamento_log(agendamento_id, dados):
    data = dados.get('data_especifica') or dados.get('dia') or dados.get('dia_semana') or ''
    return (
        f'Agendamento #{agendamento_id} - '
        f'sala: {dados.get("sala", "")}, '
        f'horario: {dados.get("horario", "")}, '
        f'data/dia: {data}'
    )


def criar_backup_sqlite_bytes():
    return backup_utils.criar_backup_sqlite_bytes(DB_PATH)


def validar_backup_sqlite(caminho):
    return backup_utils.validar_backup_sqlite(caminho)


def salvar_backup_antes_da_restauracao():
    return backup_utils.salvar_backup_antes_da_restauracao(DB_PATH, BACKUP_DIR)


def remover_arquivos_sqlite_auxiliares():
    backup_utils.remover_arquivos_sqlite_auxiliares(DB_PATH)


def limpar_backups_antigos(pasta, dias=BACKUP_RETENTION_DAYS):
    return backup_utils.limpar_backups_antigos(pasta, dias)


def salvar_backup_automatico():
    resultado = backup_utils.salvar_backup_automatico(DB_PATH, BACKUP_DIR, BACKUP_RETENTION_DAYS)
    registrar_log(
        'BACKUP',
        f'Backup automático salvo em {resultado["arquivo"]}; '
        f'tamanho={resultado["tamanho"]} bytes; '
        f'antigos_removidos={resultado["antigos_removidos"]}'
    )
    return resultado


def checar_conflito(dia, horario, sala, data_especifica='', excluir_id=None):
    conn = get_db()
    try:
        try:
            eid = int(excluir_id) if excluir_id else None
        except (ValueError, TypeError):
            eid = None

        data_especifica = (data_especifica or '').strip()
        params = [horario, sala]
        q = 'SELECT * FROM agendamentos WHERE horario=? AND sala=? AND ocupa_sala=1'

        if data_especifica:
            q += (
                ' AND (data_especifica=? '
                'OR (dia_semana=? AND (data_especifica IS NULL OR data_especifica = \'\')))'
            )
            params.extend([data_especifica, dia])
        else:
            q += (
                ' AND ('
                '(dia_semana=? AND (data_especifica IS NULL OR data_especifica = \'\')) '
                'OR (data_especifica IS NOT NULL AND data_especifica != \'\' '
                'AND data_especifica >= ? AND strftime(\'%w\', data_especifica)=?)'
                ')'
            )
            params.extend([dia, data_hoje_iso(), numero_semana_sqlite(dia)])

        if eid:
            q += ' AND CAST(id AS INTEGER)!=?'
            params.append(eid)

        r = conn.execute(q, params).fetchone()
    finally:
        conn.close()
    return dict(r) if r else None


def contar_reservas_pendentes():
    total = reservas_mod.contar_reservas_pendentes(get_db)
    if not current_user.is_authenticated or current_user.role not in ('coordenador', 'recepcao'):
        return total
    conn = get_db()
    try:
        row = conn.execute(
            """
            SELECT COUNT(*) AS total
            FROM solicitacoes_vagas
            WHERE status IN ('pendente', 'em_analise', 'atendida_parcial')
            """
        ).fetchone()
        return total + (row['total'] if row else 0)
    finally:
        conn.close()


def validar_antecedencia_minima(data_uso, horario_inicio):
    return reservas_mod.validar_antecedencia_minima(data_uso, horario_inicio)


def horarios_do_intervalo(horario_inicio, horario_fim):
    return reservas_mod.horarios_do_intervalo(horario_inicio, horario_fim, HORARIOS)


def candidatos_sala_reserva(tipo_sala):
    return reservas_mod.candidatos_sala_reserva(
        tipo_sala,
        SALAS,
        SALAS_RESERVAVEIS,
        SALAS_COM_COMPUTADOR
    )


def encontrar_sala_disponivel(data_uso, horario_inicio, horario_fim, tipo_sala):
    return reservas_mod.encontrar_sala_disponivel(
        data_uso,
        horario_inicio,
        horario_fim,
        tipo_sala,
        normalizar_data_especifica=normalizar_data_especifica,
        dia_semana_da_data=dia_semana_da_data,
        checar_conflito=checar_conflito,
        horarios=HORARIOS,
        salas=SALAS,
        salas_reservaveis=SALAS_RESERVAVEIS,
        salas_com_computador=SALAS_COM_COMPUTADOR,
    )


def sala_disponivel_para_reserva(data_uso, horario_inicio, horario_fim, sala):
    return reservas_mod.sala_disponivel_para_reserva(
        data_uso,
        horario_inicio,
        horario_fim,
        sala,
        normalizar_data_especifica=normalizar_data_especifica,
        dia_semana_da_data=dia_semana_da_data,
        checar_conflito=checar_conflito,
        horarios=HORARIOS,
        salas_reservaveis=SALAS_RESERVAVEIS,
    )


def label_status_reserva(status):
    return reservas_mod.label_status_reserva(status)

# ========================================
# REGRAS DE NEGOCIO
# ========================================

def normalize(t):
    return agendamento_utils.normalize(t)


def detect_cat(est, pac):
    return agendamento_utils.detect_cat(est, pac)


def detect_sem(t):
    return agendamento_utils.detect_sem(t)


# ========================================
# API PUBLICA
# ========================================

def pagina_erro(titulo, mensagem, status):
    try:
        return render_template(
            'error.html',
            titulo=titulo,
            mensagem=mensagem,
            status=status,
            versao=VERSAO
        ), status
    except TemplateNotFound:
        return render_template_string(
            '''
            <!doctype html>
            <html lang="pt-BR">
            <head>
              <meta charset="utf-8">
              <meta name="viewport" content="width=device-width, initial-scale=1">
              <title>{{ status }} - {{ titulo }}</title>
              <style>
                body{font-family:Arial,sans-serif;background:#f4f6f9;color:#1f2937;margin:0}
                main{max-width:680px;margin:80px auto;padding:24px}
                .box{background:#fff;border-radius:10px;padding:28px;box-shadow:0 10px 30px rgba(0,0,0,.08)}
                h1{margin:0 0 10px;font-size:28px}
                p{line-height:1.5;color:#4b5563}
                a{display:inline-block;margin-top:14px;background:#2563eb;color:#fff;text-decoration:none;padding:10px 14px;border-radius:8px}
                footer{margin-top:18px;color:#6b7280;font-size:12px}
              </style>
            </head>
            <body>
              <main>
                <div class="box">
                  <h1>{{ status }} - {{ titulo }}</h1>
                  <p>{{ mensagem }}</p>
                  <a href="{{ url_for('index') }}">Voltar ao mapa</a>
                  <footer>Versão {{ versao }}</footer>
                </div>
              </main>
            </body>
            </html>
            ''',
            titulo=titulo,
            mensagem=mensagem,
            status=status,
            versao=VERSAO
        ), status


@app.errorhandler(403)
def erro_403(e):
    return pagina_erro('Acesso negado', 'Você não tem permissão para acessar esta área.', 403)


@app.errorhandler(404)
def erro_404(e):
    return pagina_erro('Página não encontrada', 'Confira o endereço ou volte para o mapa de salas.', 404)


@app.errorhandler(500)
def erro_500(e):
    return pagina_erro('Erro interno', 'Algo saiu do esperado. Tente novamente e avise a coordenação se persistir.', 500)


@app.route('/api/versao')
def api_versao():
    return jsonify({'versao': VERSAO, 'ok': True})


# ========================================
# ROTAS DE AUTENTICACAO
# ========================================

@app.route('/login', methods=['GET', 'POST'])
@limiter.limit('5 per minute', methods=['POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('index'))
    if request.method == 'POST':
        identificador = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        conn = get_db()
        try:
            row = conn.execute(
                'SELECT * FROM usuarios WHERE username=? OR email=?',
                (identificador, identificador)
            ).fetchone()
        finally:
            conn.close()
        if row and not row['ativo']:
            flash('Usuário inativo. Procure a coordenação.')
            return render_template('login.html')
        if row and check_password_hash(row['password_hash'], password):
            user = Usuario(row['id'], row['username'], row['role'],
                           row['nome_completo'] or '', row['email'] or '', row['ativo'])
            session.permanent = True
            session['ultimo_uso'] = datetime.now(timezone.utc).isoformat()
            login_user(user)
            registrar_log('LOGIN', f'Usuário {row["username"]} fez login')
            return redirect(url_for('index'))
        flash('Usuário ou senha inválidos.')
    return render_template('login.html')


@app.route('/recuperar-senha', methods=['GET', 'POST'])
@limiter.limit('5 per minute', methods=['POST'])
def recuperar_senha():
    if current_user.is_authenticated:
        return redirect(url_for('index'))

    if request.method == 'POST':
        identificador = request.form.get('identificador', '').strip()
        conn = get_db()
        try:
            row = conn.execute(
                "SELECT * FROM usuarios WHERE ativo=1 AND (username=? OR email=?)",
                (identificador, identificador)
            ).fetchone()
            if row and row['email']:
                token = criar_token_email(conn, row['id'], 'reset', horas=72)
                conn.commit()
                link = url_absoluta('redefinir_senha', token=token)
                enviar_email(
                    row['email'],
                    'Redefinição de senha',
                    corpo_email_padrao(
                        f'Olá, {nome_exibicao_usuario(row)}!',
                        'Recebemos uma solicitação para redefinir sua senha de acesso ao Mapa de Sala.',
                        observacao=(
                            'Clique no link abaixo para cadastrar uma nova senha:\n\n'
                            f'{link}\n\n'
                            'Este link permanecerá válido por 72 horas.\n'
                            'Se você não realizou essa solicitação, ignore este e-mail. Sua senha atual permanecerá válida.'
                        )
                    )
                )
        finally:
            conn.close()

        flash('Se o usuário ou e-mail existir no sistema, enviaremos um link de recuperação.', 'success')
        return redirect(url_for('login'))

    return render_template('recuperar_senha.html')


@app.route('/definir-senha/<token>', methods=['GET', 'POST'])
@limiter.limit('10 per minute', methods=['POST'])
def redefinir_senha(token):
    row_reset = buscar_token_email(token, 'reset')
    row_convite = buscar_token_email(token, 'convite') if not row_reset else None
    row = row_reset or row_convite
    tipo = 'convite' if row_convite else 'reset'
    if not row:
        flash('Link inválido ou expirado. Peça um novo link.', 'error')
        return redirect(url_for('login'))

    if request.method == 'POST':
        nova_senha = request.form.get('nova_senha', '').strip()
        confirmar = request.form.get('confirmar_senha', '').strip()
        if len(nova_senha) < 8:
            flash('A senha deve ter no mínimo 8 caracteres.', 'error')
            return redirect(url_for('redefinir_senha', token=token))
        if nova_senha != confirmar:
            flash('As senhas não coincidem.', 'error')
            return redirect(url_for('redefinir_senha', token=token))

        conn = get_db()
        try:
            conn.execute(
                'UPDATE usuarios SET password_hash=?, ativo=1 WHERE id=?',
                (generate_password_hash(nova_senha), row['usuario_id'])
            )
            marcar_token_usado(conn, row['id'])
            conn.commit()
        finally:
            conn.close()

        registrar_log('DEFINIR_SENHA_EMAIL', f'Usuário {row["username"]} definiu senha por link de {tipo}')
        try:
            notificar_senha_alterada_email(row, 'reset')
        except Exception as exc:
            app.logger.warning('Falha ao enviar aviso de senha alterada: %s', exc)
        flash('Senha definida com sucesso. Faça login para entrar.', 'success')
        return redirect(url_for('login'))

    return render_template(
        'redefinir_senha.html',
        token=token,
        titulo='Criar senha' if tipo == 'convite' else 'Recuperar senha',
        usuario=row['username']
    )


@app.route('/logout')
@login_required
def logout():
    registrar_log('LOGOUT', f'Usuário {current_user.username} saiu')
    logout_user()
    return redirect(url_for('login'))


# ========================================
# ROTAS DE PAGINAS
# ========================================

@app.route('/')
@login_required
def index():
    if current_user.role == 'aluno':
        return redirect(url_for('meus_agendamentos'))
    if current_user.role == 'professor':
        return redirect(url_for('minha_supervisao'))
    if current_user.role == 'recepcao':
        return redirect(url_for('painel_recepcao'))
    return renderizar_mapa_sala()


@app.route('/mapa')
@login_required
def mapa_sala():
    if current_user.role == 'aluno':
        return redirect(url_for('meus_agendamentos'))
    if current_user.role == 'professor':
        return redirect(url_for('minha_supervisao'))
    return renderizar_mapa_sala()


def renderizar_mapa_sala():
    with db_connection() as conn:
        professores = listar_professores_ativos(conn)
    return render_template(
        'index.html',
        salas=SALAS,
        horarios=HORARIOS,
        categorias=CATEGORIAS,
        categorias_operacionais=CATEGORIAS_OPERACIONAIS,
        dias=DIAS,
        professores=professores,
        usuario=current_user.username,
        papel=current_user.role
    )


def coletar_painel_coordenacao():
    hoje = data_hoje_iso()
    with db_connection() as conn:
        return relatorios.coletar_painel_coordenacao(conn, hoje, DIAS, DIAS_PT)


@app.route('/painel-coordenacao')
@login_required
@requer_papel_page('coordenador', 'somente_leitura')
def painel_coordenacao():
    dados = coletar_painel_coordenacao()
    return render_template(
        'painel_coordenacao.html',
        usuario=current_user.username,
        papel=current_user.role,
        papel_label=PAPEIS_LABEL.get(current_user.role, current_user.role),
        dados=dados,
        versao=VERSAO
    )


@app.route('/painel')
@login_required
@requer_papel_page('coordenador', 'recepcao', 'somente_leitura')
def painel_recepcao():
    hoje = data_hoje_iso()
    weekday = datetime.now().weekday()
    dia_hoje = DIAS[weekday] if weekday < len(DIAS) else ''

    conn = get_db()
    try:
        pendentes_sala = conn.execute(
            """
            SELECT *
            FROM reservas
            WHERE tipo='sala' AND status='pendente'
            ORDER BY data_uso, horario_inicio
            LIMIT 6
            """
        ).fetchall()
        pendentes_instrumento = conn.execute(
            """
            SELECT *
            FROM reservas
            WHERE tipo='instrumento' AND status='pendente'
            ORDER BY data_uso, horario_inicio
            LIMIT 6
            """
        ).fetchall()
        instrumentos_ativos = conn.execute(
            """
            SELECT *
            FROM reservas
            WHERE tipo='instrumento'
              AND status IN ('aprovada', 'separado')
              AND data_uso>=?
            ORDER BY data_uso, horario_inicio
            LIMIT 8
            """,
            (hoje,)
        ).fetchall()
        tarefas_manuais = conn.execute(
            """
            SELECT *
            FROM tarefas_painel
            ORDER BY created_at DESC
            LIMIT 12
            """
        ).fetchall()
        solicitacoes_vagas_pendentes = conn.execute(
            """
            SELECT *
            FROM solicitacoes_vagas
            WHERE status IN ('pendente', 'em_analise', 'atendida_parcial')
            ORDER BY created_at ASC
            LIMIT 8
            """
        ).fetchall()
        horarios_abertos_rows = conn.execute(
            """
            SELECT *
            FROM agendamentos
            WHERE (data_especifica IS NULL OR data_especifica='')
              AND TRIM(COALESCE(paciente, ''))=''
              AND (
                categoria='MARCAR'
                OR triagem=1
              )
            ORDER BY
              CASE dia_semana
                WHEN 'SEGUNDA' THEN 1
                WHEN 'TERÇA' THEN 2
                WHEN 'QUARTA' THEN 3
                WHEN 'QUINTA' THEN 4
                WHEN 'SEXTA' THEN 5
                ELSE 6
              END,
              horario,
              sala
            LIMIT 8
            """
        ).fetchall()
        total_horarios_abertos = conn.execute(
            """
            SELECT COUNT(*) AS total
            FROM agendamentos
            WHERE (data_especifica IS NULL OR data_especifica='')
              AND TRIM(COALESCE(paciente, ''))=''
              AND (
                categoria='MARCAR'
                OR triagem=1
              )
            """
        ).fetchone()['total']

        if dia_hoje:
            ag_hoje = conn.execute(
                """
                SELECT *
                FROM agendamentos
                WHERE ocupa_sala=1
                  AND (
                    data_especifica=?
                    OR (dia_semana=? AND (data_especifica IS NULL OR data_especifica=''))
                  )
                ORDER BY horario, sala
                LIMIT 10
                """,
                (hoje, dia_hoje)
            ).fetchall()
        else:
            ag_hoje = conn.execute(
                """
                SELECT *
                FROM agendamentos
                WHERE ocupa_sala=1 AND data_especifica=?
                ORDER BY horario, sala
                LIMIT 10
                """,
                (hoje,)
            ).fetchall()
    finally:
        conn.close()

    def preparar_linha_reserva(row):
        r = dict(row)
        r['status_label'] = label_status_reserva(r.get('status'))
        try:
            r['data_label'] = datetime.strptime(r['data_uso'], '%Y-%m-%d').strftime('%d/%m/%Y')
        except (ValueError, TypeError):
            r['data_label'] = r.get('data_uso') or ''
        return r

    checklist_auto = []
    if pendentes_sala:
        checklist_auto.append({
            'titulo': 'Responder reservas de sala',
            'detalhe': f'{len(pendentes_sala)} solicitação(ões) aguardando aprovação.',
            'url': url_for('reservas')
        })
    if pendentes_instrumento:
        checklist_auto.append({
            'titulo': 'Responder pedidos de instrumentos',
            'detalhe': f'{len(pendentes_instrumento)} pedido(s) aguardando aprovação.',
            'url': url_for('reservas')
        })
    instrumentos_separados = [r for r in instrumentos_ativos if r['status'] == 'separado']
    if instrumentos_separados:
        checklist_auto.append({
            'titulo': 'Guardar instrumentos separados',
            'detalhe': f'{len(instrumentos_separados)} instrumento(s) separado(s) aguardando guarda.',
            'url': url_for('reservas')
        })
    if total_horarios_abertos and current_user.role == 'coordenador':
        checklist_auto.append({
            'titulo': 'Ver horários abertos sem paciente',
            'detalhe': f'{total_horarios_abertos} horário(s) para marcar paciente ou triagem.',
            'url': url_for('horarios_abertos')
        })
    if solicitacoes_vagas_pendentes:
        total_vagas = sum((r['vagas_paciente'] or 0) + (r['vagas_triagem'] or 0) for r in solicitacoes_vagas_pendentes)
        checklist_auto.append({
            'titulo': 'Responder solicitações de vagas',
            'detalhe': f'{len(solicitacoes_vagas_pendentes)} pedido(s), somando {total_vagas} vaga(s).',
            'url': url_for('reservas') + '#pedidos-vagas'
        })

    return render_template(
        'painel_recepcao.html',
        usuario=current_user.username,
        papel=current_user.role,
        papel_label=PAPEIS_LABEL.get(current_user.role, current_user.role),
        hoje_label=datetime.now().strftime('%d/%m/%Y'),
        dia_hoje_label=DIAS_PT.get(dia_hoje, 'Hoje'),
        pendentes_sala=[preparar_linha_reserva(r) for r in pendentes_sala],
        pendentes_instrumento=[preparar_linha_reserva(r) for r in pendentes_instrumento],
        instrumentos_ativos=[preparar_linha_reserva(r) for r in instrumentos_ativos],
        agendamentos_hoje=[dict(r) for r in ag_hoje],
        tarefas_manuais=[dict(r) for r in tarefas_manuais],
        checklist_auto=checklist_auto,
        horarios_abertos_preview=[dict(r) for r in horarios_abertos_rows],
        total_horarios_abertos=total_horarios_abertos,
        solicitacoes_vagas=[preparar_solicitacao_vaga(r) for r in solicitacoes_vagas_pendentes],
        versao=VERSAO,
    )


@app.route('/painel/tarefas', methods=['POST'])
@login_required
@requer_papel_page('coordenador', 'recepcao')
def criar_tarefa_painel():
    titulo = request.form.get('titulo', '').strip()
    detalhe = request.form.get('detalhe', '').strip()
    if not titulo:
        flash('Informe a tarefa antes de adicionar.', 'error')
        return redirect(url_for('painel_recepcao'))

    conn = get_db()
    try:
        conn.execute(
            "INSERT INTO tarefas_painel(titulo, detalhe, criado_por) VALUES(?,?,?)",
            (titulo, detalhe, current_user.username)
        )
        conn.commit()
    finally:
        conn.close()

    registrar_log('CRIAR_TAREFA_PAINEL', f'{current_user.username} criou tarefa: {titulo}')
    flash('Tarefa adicionada aos afazeres.', 'success')
    return redirect(url_for('afazeres_recepcao'))


@app.route('/minha-supervisao/solicitacoes-vagas', methods=['POST'])
@login_required
@requer_papel_page('professor')
def criar_solicitacao_vagas():
    aluno_id_raw = request.form.get('aluno_id', '').strip()
    observacao = request.form.get('observacao', '').strip()

    try:
        aluno_id = int(aluno_id_raw)
        vagas_paciente = max(0, int(request.form.get('vagas_paciente') or 0))
        vagas_triagem = max(0, int(request.form.get('vagas_triagem') or 0))
    except (TypeError, ValueError):
        flash('Revise os números da solicitação de vagas.', 'error')
        return redirect(url_for('minha_supervisao'))

    if vagas_paciente + vagas_triagem <= 0:
        flash('Informe pelo menos uma vaga de paciente ou de triagem.', 'error')
        return redirect(url_for('minha_supervisao'))
    if vagas_paciente > 20 or vagas_triagem > 20:
        flash('Use até 20 vagas por tipo em cada solicitação.', 'error')
        return redirect(url_for('minha_supervisao'))

    conn = get_db()
    try:
        aluno = conn.execute(
            """
            SELECT id, username, nome_completo
            FROM usuarios
            WHERE id=? AND role='aluno' AND ativo=1 AND supervisor_id=?
            """,
            (aluno_id, current_user.id)
        ).fetchone()
        if not aluno:
            flash('Aluno não encontrado na sua supervisão.', 'error')
            return redirect(url_for('minha_supervisao'))

        pedido_aberto = conn.execute(
            """
            SELECT id
            FROM solicitacoes_vagas
            WHERE professor_id=?
              AND aluno_id=?
              AND status IN ('pendente', 'em_analise', 'atendida_parcial')
            ORDER BY created_at DESC
            LIMIT 1
            """,
            (current_user.id, aluno_id)
        ).fetchone()
        if pedido_aberto:
            flash('Este aluno já tem um pedido em aberto. Acompanhe o retorno antes de enviar outro.', 'error')
            return redirect(url_for('minha_supervisao') + '#solicitacoes-vagas')

        professor_nome = current_user.nome_completo or current_user.username
        aluno_nome = aluno['nome_completo'] or aluno['username']
        conn.execute(
            """
            INSERT INTO solicitacoes_vagas(
                professor_id, professor_nome, aluno_id, aluno_nome,
                vagas_paciente, vagas_triagem, observacao
            ) VALUES(?,?,?,?,?,?,?)
            """,
            (current_user.id, professor_nome, aluno['id'], aluno_nome, vagas_paciente, vagas_triagem, observacao)
        )
        conn.commit()
    finally:
        conn.close()

    registrar_log(
        'SOLICITAR_VAGAS',
        f'{current_user.username} solicitou {vagas_paciente} paciente(s) e {vagas_triagem} triagem(ns) para {aluno_nome}'
    )
    flash('Solicitação enviada para recepção/coordenação.', 'success')
    return redirect(url_for('minha_supervisao') + '#solicitacoes-vagas')


@app.route('/solicitacoes-vagas/<int:solicitacao_id>/status', methods=['POST'])
@login_required
@requer_papel_page('coordenador', 'recepcao')
def atualizar_solicitacao_vagas(solicitacao_id):
    status = request.form.get('status', '').strip()
    resposta = request.form.get('resposta', '').strip()
    proximo = request.form.get('next') or request.referrer or (url_for('painel_recepcao') + '#solicitacoes-vagas')
    if not str(proximo).startswith('/'):
        proximo = url_for('painel_recepcao') + '#solicitacoes-vagas'
    if status not in STATUS_SOLICITACAO_VAGA:
        flash('Status inválido para solicitação de vagas.', 'error')
        return redirect(proximo)

    try:
        liberadas_paciente = max(0, int(request.form.get('vagas_paciente_liberadas') or 0))
        liberadas_triagem = max(0, int(request.form.get('vagas_triagem_liberadas') or 0))
    except (TypeError, ValueError):
        flash('Informe números válidos para as vagas liberadas.', 'error')
        return redirect(proximo)

    conn = get_db()
    try:
        row = conn.execute(
            """
            SELECT id, aluno_nome, vagas_paciente, vagas_triagem
            FROM solicitacoes_vagas
            WHERE id=?
            """,
            (solicitacao_id,)
        ).fetchone()
        if not row:
            flash('Solicitação de vagas não encontrada.', 'error')
            return redirect(proximo)

        vagas_paciente = int(row['vagas_paciente'] or 0)
        vagas_triagem = int(row['vagas_triagem'] or 0)
        if status == 'atendida':
            liberadas_paciente = vagas_paciente
            liberadas_triagem = vagas_triagem
        elif status == 'recusada':
            liberadas_paciente = 0
            liberadas_triagem = 0

        if liberadas_paciente > vagas_paciente or liberadas_triagem > vagas_triagem:
            flash('O número liberado não pode ser maior do que o pedido pelo professor.', 'error')
            return redirect(proximo)

        if status == 'atendida_parcial':
            if liberadas_paciente + liberadas_triagem <= 0:
                flash('Para atender em parte, informe pelo menos uma vaga liberada.', 'error')
                return redirect(proximo)
            if liberadas_paciente == vagas_paciente and liberadas_triagem == vagas_triagem:
                status = 'atendida'

        conn.execute(
            """
            UPDATE solicitacoes_vagas
            SET status=?,
                resposta=?,
                vagas_paciente_liberadas=?,
                vagas_triagem_liberadas=?,
                analisado_por=?,
                updated_at=CURRENT_TIMESTAMP
            WHERE id=?
            """,
            (status, resposta, liberadas_paciente, liberadas_triagem, current_user.username, solicitacao_id)
        )
        conn.commit()
    finally:
        conn.close()

    registrar_log(
        'ATUALIZAR_SOLICITACAO_VAGAS',
        f'{current_user.username} marcou solicitação #{solicitacao_id} como {status}'
    )
    flash('Solicitação de vagas atualizada.', 'success')
    return redirect(proximo)


@app.route('/painel/tarefas/<int:tid>/concluir', methods=['POST'])
@login_required
@requer_papel_page('coordenador', 'recepcao')
def concluir_tarefa_painel(tid):
    tarefa = None
    conn = get_db()
    try:
        tarefa = conn.execute('SELECT * FROM tarefas_painel WHERE id=?', (tid,)).fetchone()
        if tarefa:
            conn.execute('DELETE FROM tarefas_painel WHERE id=?', (tid,))
            conn.commit()
    finally:
        conn.close()

    if tarefa:
        registrar_log('CONCLUIR_TAREFA_PAINEL', f'{current_user.username} concluiu tarefa: {tarefa["titulo"]}')
        flash('Tarefa concluída e removida dos afazeres.', 'success')
    else:
        flash('Tarefa não encontrada.', 'error')
    return redirect(url_for('afazeres_recepcao'))


@app.route('/afazeres')
@login_required
@requer_papel_page('coordenador', 'recepcao', 'somente_leitura')
def afazeres_recepcao():
    conn = get_db()
    try:
        tarefas = conn.execute(
            """
            SELECT *
            FROM tarefas_painel
            ORDER BY created_at DESC
            """
        ).fetchall()
    finally:
        conn.close()

    return render_template(
        'afazeres.html',
        usuario=current_user.username,
        papel=current_user.role,
        papel_label=PAPEIS_LABEL.get(current_user.role, current_user.role),
        tarefas=[dict(t) for t in tarefas],
    )


def row_coord_agendamento(conn, agendamento_id):
    return conn.execute(
        """
        SELECT ca.*,
               ch.data_disponivel,
               ch.horario_inicio,
               ch.horario_fim,
               ch.local,
               ch.coordenador_nome,
               u.email AS aluno_email
        FROM coordenacao_agendamentos ca
        JOIN coordenacao_horarios ch ON ch.id = ca.horario_id
        LEFT JOIN usuarios u ON u.id = ca.aluno_id
        WHERE ca.id=?
        """,
        (agendamento_id,)
    ).fetchone()


@app.route('/coordenacao')
@login_required
@requer_papel_page('coordenador', 'aluno', 'somente_leitura')
def coordenacao_page():
    hoje = data_hoje_iso()
    conn = get_db()
    try:
        horarios_rows = conn.execute(
            """
            SELECT ch.*,
                   ca.id AS agendamento_id,
                   ca.aluno_nome,
                   ca.assunto,
                   ca.status AS agendamento_status
            FROM coordenacao_horarios ch
            LEFT JOIN coordenacao_agendamentos ca
              ON ca.horario_id = ch.id
             AND ca.status='confirmado'
            WHERE ch.data_disponivel >= ?
              AND ch.ativo=1
            ORDER BY ch.data_disponivel, ch.horario_inicio
            """,
            (hoje,)
        ).fetchall()
        meus_rows = conn.execute(
            """
            SELECT ca.*,
                   ch.data_disponivel,
                   ch.horario_inicio,
                   ch.horario_fim,
                   ch.local,
                   ch.coordenador_nome
            FROM coordenacao_agendamentos ca
            JOIN coordenacao_horarios ch ON ch.id = ca.horario_id
            WHERE ca.aluno_id=?
              AND ch.data_disponivel >= ?
            ORDER BY ch.data_disponivel, ch.horario_inicio
            """,
            (current_user.id, hoje)
        ).fetchall() if current_user.role == 'aluno' else []
        agendados_rows = conn.execute(
            """
            SELECT ca.*,
                   ch.data_disponivel,
                   ch.horario_inicio,
                   ch.horario_fim,
                   ch.local,
                   ch.coordenador_nome,
                   u.email AS aluno_email
            FROM coordenacao_agendamentos ca
            JOIN coordenacao_horarios ch ON ch.id = ca.horario_id
            LEFT JOIN usuarios u ON u.id = ca.aluno_id
            WHERE ch.data_disponivel >= ?
            ORDER BY ch.data_disponivel, ch.horario_inicio
            """,
            (hoje,)
        ).fetchall() if current_user.role in ('coordenador', 'somente_leitura') else []
    finally:
        conn.close()

    horarios = []
    for row in horarios_rows:
        item = dict(row)
        item['data_label'] = formatar_data_iso(item.get('data_disponivel'))
        item['ocupado'] = bool(item.get('agendamento_id'))
        horarios.append(item)

    meus_agendamentos = []
    for row in meus_rows:
        item = dict(row)
        item['data_label'] = formatar_data_iso(item.get('data_disponivel'))
        meus_agendamentos.append(item)

    agendados = []
    for row in agendados_rows:
        item = dict(row)
        item['data_label'] = formatar_data_iso(item.get('data_disponivel'))
        agendados.append(item)

    return render_template(
        'coordenacao.html',
        usuario=current_user.username,
        papel=current_user.role,
        papel_label=PAPEIS_LABEL.get(current_user.role, current_user.role),
        horarios=horarios,
        meus_agendamentos=meus_agendamentos,
        agendados=agendados,
        horarios_padrao=HORARIOS,
    )


@app.route('/coordenacao/horarios', methods=['POST'])
@login_required
@requer_papel_page('coordenador')
def criar_horario_coordenacao():
    data_disponivel = request.form.get('data_disponivel', '').strip()
    horario_inicio = request.form.get('horario_inicio', '').strip()
    horario_fim = request.form.get('horario_fim', '').strip()
    local = request.form.get('local', '').strip()
    observacao = request.form.get('observacao', '').strip()

    data_disponivel, erro_data = normalizar_data_especifica(data_disponivel)
    if erro_data:
        flash(erro_data, 'error')
        return redirect(url_for('coordenacao_page'))
    if data_disponivel < data_hoje_iso():
        flash('Escolha uma data de hoje em diante.', 'error')
        return redirect(url_for('coordenacao_page'))
    if horario_inicio not in HORARIOS:
        flash('Escolha um horário inicial válido.', 'error')
        return redirect(url_for('coordenacao_page'))
    if horario_fim and horario_fim not in HORARIOS:
        flash('Escolha um horário final válido.', 'error')
        return redirect(url_for('coordenacao_page'))
    if horario_fim and HORARIOS.index(horario_fim) <= HORARIOS.index(horario_inicio):
        flash('O horário final deve ser depois do inicial.', 'error')
        return redirect(url_for('coordenacao_page'))

    coordenador_nome = current_user.nome_completo or current_user.username
    conn = get_db()
    try:
        existe = conn.execute(
            """
            SELECT id
            FROM coordenacao_horarios
            WHERE data_disponivel=? AND horario_inicio=? AND ativo=1
            """,
            (data_disponivel, horario_inicio)
        ).fetchone()
        if existe:
            flash('Já existe um horário publicado nessa data e horário.', 'error')
            return redirect(url_for('coordenacao_page'))
        conn.execute(
            """
            INSERT INTO coordenacao_horarios(
                coordenador_id, coordenador_nome, data_disponivel,
                horario_inicio, horario_fim, local, observacao
            ) VALUES(?,?,?,?,?,?,?)
            """,
            (current_user.id, coordenador_nome, data_disponivel, horario_inicio, horario_fim, local, observacao)
        )
        conn.commit()
    finally:
        conn.close()

    registrar_log('CRIAR_HORARIO_COORDENACAO', f'{current_user.username} abriu {data_disponivel} {horario_inicio}')
    flash('Horário disponibilizado para os alunos.', 'success')
    return redirect(url_for('coordenacao_page'))


@app.route('/coordenacao/horarios/<int:horario_id>/desativar', methods=['POST'])
@login_required
@requer_papel_page('coordenador')
def desativar_horario_coordenacao(horario_id):
    conn = get_db()
    try:
        ocupado = conn.execute(
            "SELECT id FROM coordenacao_agendamentos WHERE horario_id=? AND status='confirmado'",
            (horario_id,)
        ).fetchone()
        if ocupado:
            flash('Esse horário já tem aluno marcado. Cancele o agendamento antes de remover o horário.', 'error')
            return redirect(url_for('coordenacao_page'))
        conn.execute('UPDATE coordenacao_horarios SET ativo=0, updated_at=CURRENT_TIMESTAMP WHERE id=?', (horario_id,))
        conn.commit()
    finally:
        conn.close()

    registrar_log('DESATIVAR_HORARIO_COORDENACAO', f'{current_user.username} removeu horário #{horario_id}')
    flash('Horário removido da lista de disponibilidade.', 'success')
    return redirect(url_for('coordenacao_page'))


@app.route('/coordenacao/agendar/<int:horario_id>', methods=['POST'])
@login_required
@requer_papel_page('aluno')
def agendar_horario_coordenacao(horario_id):
    assunto = request.form.get('assunto', '').strip()
    observacao = request.form.get('observacao', '').strip()
    if not assunto:
        flash('Informe rapidamente o assunto da conversa.', 'error')
        return redirect(url_for('coordenacao_page'))

    aluno_nome = current_user.nome_completo or current_user.username
    conn = get_db()
    try:
        horario = conn.execute(
            """
            SELECT *
            FROM coordenacao_horarios
            WHERE id=? AND ativo=1 AND data_disponivel>=?
            """,
            (horario_id, data_hoje_iso())
        ).fetchone()
        if not horario:
            flash('Esse horário não está mais disponível.', 'error')
            return redirect(url_for('coordenacao_page'))
        ocupado = conn.execute(
            "SELECT id FROM coordenacao_agendamentos WHERE horario_id=? AND status='confirmado'",
            (horario_id,)
        ).fetchone()
        if ocupado:
            flash('Esse horário acabou de ser reservado por outro aluno.', 'error')
            return redirect(url_for('coordenacao_page'))

        cur = conn.execute(
            """
            INSERT INTO coordenacao_agendamentos(horario_id, aluno_id, aluno_nome, assunto, observacao)
            VALUES(?,?,?,?,?)
            """,
            (horario_id, current_user.id, aluno_nome, assunto, observacao)
        )
        conn.commit()
        agendamento = row_coord_agendamento(conn, cur.lastrowid)
    finally:
        conn.close()

    if agendamento:
        notificar_coord_agendamento_email(dict(agendamento), 'criado')
    registrar_log('AGENDAR_COORDENACAO', f'{current_user.username} reservou horário #{horario_id}')
    flash('Horário com a coordenação confirmado.', 'success')
    return redirect(url_for('coordenacao_page'))


@app.route('/coordenacao/agendamentos/<int:agendamento_id>/cancelar', methods=['POST'])
@login_required
@requer_papel_page('coordenador', 'aluno')
def cancelar_agendamento_coordenacao(agendamento_id):
    resposta = request.form.get('resposta', '').strip()
    conn = get_db()
    try:
        agendamento = row_coord_agendamento(conn, agendamento_id)
        if not agendamento:
            flash('Agendamento não encontrado.', 'error')
            return redirect(url_for('coordenacao_page'))
        if current_user.role == 'aluno' and agendamento['aluno_id'] != current_user.id:
            flash('Você não pode cancelar um horário de outro aluno.', 'error')
            return redirect(url_for('coordenacao_page'))
        conn.execute(
            """
            UPDATE coordenacao_agendamentos
            SET status='cancelado', resposta=?, updated_at=CURRENT_TIMESTAMP
            WHERE id=?
            """,
            (resposta, agendamento_id)
        )
        conn.commit()
        agendamento_atualizado = row_coord_agendamento(conn, agendamento_id)
    finally:
        conn.close()

    if agendamento_atualizado:
        notificar_coord_agendamento_email(dict(agendamento_atualizado), 'cancelado', resposta=resposta)
    registrar_log('CANCELAR_COORDENACAO', f'{current_user.username} cancelou horário de coordenação #{agendamento_id}')
    flash('Horário com a coordenação cancelado.', 'success')
    return redirect(url_for('coordenacao_page'))


@app.route('/sobre')
@login_required
def sobre():
    return redirect('/informacoes#sobre')


@app.route('/termo-uso')
@login_required
def termo_uso():
    return redirect('/informacoes#termo')


@app.route('/informacoes')
@login_required
def informacoes():
    return render_template(
        'informacoes.html',
        usuario=current_user.username,
        papel=current_user.role,
        papel_label=PAPEIS_LABEL.get(current_user.role, current_user.role),
        versao=VERSAO
    )


@app.route('/ajuda')
@app.route('/ajuda/<topico>')
@login_required
def ajuda(topico='recepcao'):
    topicos_validos = ('recepcao', 'coordenacao', 'professor', 'aluno', 'backup')
    if topico not in topicos_validos:
        if current_user.role == 'professor':
            topico = 'professor'
        elif current_user.role == 'aluno':
            topico = 'aluno'
        elif current_user.role in ('coordenador', 'somente_leitura'):
            topico = 'coordenacao'
        else:
            topico = 'recepcao'
    return render_template(
        'ajuda.html',
        usuario=current_user.username,
        papel=current_user.role,
        papel_label=PAPEIS_LABEL.get(current_user.role, current_user.role),
        topico=topico,
        versao=VERSAO
    )


@app.route('/horarios-abertos')
@login_required
@requer_papel_page('coordenador', 'recepcao', 'somente_leitura')
def horarios_abertos():
    conn = get_db()
    try:
        rows = conn.execute(
            """
            SELECT *
            FROM agendamentos
            WHERE (data_especifica IS NULL OR data_especifica='')
              AND TRIM(COALESCE(paciente, ''))=''
              AND (
                categoria='MARCAR'
                OR triagem=1
              )
            ORDER BY
              CASE dia_semana
                WHEN 'SEGUNDA' THEN 1
                WHEN 'TERÇA' THEN 2
                WHEN 'QUARTA' THEN 3
                WHEN 'QUINTA' THEN 4
                WHEN 'SEXTA' THEN 5
                ELSE 6
              END,
              horario,
              sala
            """
        ).fetchall()
    finally:
        conn.close()

    horarios_por_dia = []
    for dia in DIAS:
        itens = []
        for row in rows:
            ag = dict(row)
            if ag['dia_semana'] != dia:
                continue
            ag['dia_label'] = DIAS_PT.get(dia, dia.title())
            triagem_livre = int(ag.get('triagem') or 0)
            ag['tipo_codigo'] = 'triagem' if triagem_livre else 'marcar'
            ag['tipo_abertura'] = 'Triagem livre' if triagem_livre else 'Aberto para paciente'
            ag['acao_recepcao'] = 'Procurar paciente para triagem' if triagem_livre else 'Marcar paciente'
            ag['aviso'] = (
                'Horário reservado para triagem, mas ainda sem paciente marcado.'
                if triagem_livre
                else 'Horário aberto para a recepção marcar paciente.'
            )
            itens.append(ag)
        horarios_por_dia.append({'dia': dia, 'label': DIAS_PT.get(dia, dia.title()), 'itens': itens})

    total_marcar = sum(
        1 for dia in horarios_por_dia for ag in dia['itens']
        if ag.get('tipo_codigo') == 'marcar'
    )
    total_triagem = sum(
        1 for dia in horarios_por_dia for ag in dia['itens']
        if ag.get('tipo_codigo') == 'triagem'
    )

    return render_template(
        'horarios_abertos.html',
        usuario=current_user.username,
        papel=current_user.role,
        papel_label=PAPEIS_LABEL.get(current_user.role, current_user.role),
        horarios_por_dia=horarios_por_dia,
        total=sum(len(d['itens']) for d in horarios_por_dia),
        total_marcar=total_marcar,
        total_triagem=total_triagem
    )


@app.route('/minha-supervisao')
@login_required
@requer_papel_page('professor')
def minha_supervisao():
    hoje = data_hoje_iso()
    conn = get_db()
    try:
        alunos_rows = conn.execute(
            """
            SELECT id, username, nome_completo, email
            FROM usuarios
            WHERE role='aluno'
              AND ativo=1
              AND supervisor_id=?
            ORDER BY COALESCE(NULLIF(nome_completo, ''), username)
            """,
            (current_user.id,)
        ).fetchall()

        alunos_ids = [row['id'] for row in alunos_rows]
        alunos_usernames = [row['username'] for row in alunos_rows]
        ag_rows = []
        if alunos_ids:
            placeholders_ids = ','.join('?' for _ in alunos_ids)
            placeholders_names = ','.join('?' for _ in alunos_usernames)
            ag_rows = conn.execute(
                f"""
                SELECT *
                FROM agendamentos
                WHERE (
                    usuario_id IN ({placeholders_ids})
                    OR (usuario_id IS NULL AND estagiario IN ({placeholders_names}))
                  )
                  AND (
                    data_especifica IS NULL
                    OR data_especifica = ''
                    OR data_especifica >= ?
                  )
                ORDER BY
                  CASE dia_semana
                    WHEN 'SEGUNDA' THEN 1
                    WHEN 'TERÇA' THEN 2
                    WHEN 'QUARTA' THEN 3
                    WHEN 'QUINTA' THEN 4
                    WHEN 'SEXTA' THEN 5
                    ELSE 6
                  END,
                  horario,
                  sala,
                  data_especifica
                """,
                (*alunos_ids, *alunos_usernames, hoje)
            ).fetchall()
        solicitacoes_rows = conn.execute(
            """
            SELECT *
            FROM solicitacoes_vagas
            WHERE professor_id=?
            ORDER BY created_at DESC
            LIMIT 30
            """,
            (current_user.id,)
        ).fetchall()
    finally:
        conn.close()

    ag_por_aluno = {}
    for row in ag_rows:
        ag = dict(row)
        chave = ag.get('usuario_id') or ag.get('estagiario')
        ag['dia_visual'] = ag['dia_semana']
        if ag.get('data_especifica'):
            try:
                ag['dia_visual'] = dia_semana_da_data(ag['data_especifica'])
            except ValueError:
                ag['dia_visual'] = ag['dia_semana']
        ag['dia_label'] = DIAS_PT.get(ag['dia_visual'], ag['dia_visual'].title())
        ag['tem_paciente'] = bool((ag.get('paciente') or '').strip())
        ag['eh_fixo'] = not bool(ag.get('data_especifica'))
        ag['eh_triagem'] = bool(int(ag.get('triagem') or 0))
        ag['tipo'] = 'Pontual' if ag.get('data_especifica') else 'Fixo semanal'
        ag['data_label'] = ''
        if ag.get('data_especifica'):
            try:
                ag['data_label'] = datetime.strptime(ag['data_especifica'], '%Y-%m-%d').strftime('%d/%m/%Y')
            except ValueError:
                ag['data_label'] = ag['data_especifica']
        ag_por_aluno.setdefault(chave, []).append(ag)

    alunos = []
    total_fixos = 0
    total_pacientes = 0
    total_triagens = 0
    total_abertos = 0
    total_alunos_sem_paciente = 0
    solicitacoes_por_aluno = {}
    for row in solicitacoes_rows:
        sol = preparar_solicitacao_vaga(row)
        solicitacoes_por_aluno.setdefault(sol.get('aluno_id'), []).append(sol)

    for aluno_row in alunos_rows:
        aluno = dict(aluno_row)
        itens = ag_por_aluno.get(aluno['id'], []) + ag_por_aluno.get(aluno['username'], [])
        fixos = [ag for ag in itens if ag['eh_fixo']]
        pacientes = [ag for ag in itens if ag['tem_paciente']]
        triagens = [ag for ag in itens if ag['eh_triagem']]
        abertos = []

        for ag in fixos:
            if ag['tem_paciente']:
                continue
            categoria_upper = (ag.get('categoria') or '').upper()
            if ag['eh_triagem']:
                ag['status_fixo'] = 'Triagem aberta'
                ag['status_slug'] = 'aguardando'
                ag['status_descricao'] = 'Horário reservado para triagem, ainda sem paciente marcado.'
                abertos.append(ag)
            elif categoria_upper == 'MARCAR':
                ag['status_fixo'] = 'Aberto para paciente'
                ag['status_slug'] = 'aberto'
                ag['status_descricao'] = 'Horário liberado para a recepção marcar paciente.'
                abertos.append(ag)
            else:
                ag['status_fixo'] = 'Reservado sem paciente'
                ag['status_slug'] = 'livre'
                ag['status_descricao'] = 'Horário fixo reservado, mas sem paciente marcado.'

        semana = []
        for dia in DIAS:
            dia_itens = [ag for ag in itens if ag.get('dia_visual') == dia]
            dia_itens.sort(key=lambda ag: (HORARIOS.index(ag['horario']) if ag.get('horario') in HORARIOS else 99, ag.get('sala') or ''))
            semana.append({
                'dia': dia,
                'label': DIAS_PT.get(dia, dia.title()),
                'itens': dia_itens
            })

        total_fixos += len(fixos)
        total_pacientes += len(pacientes)
        total_triagens += len(triagens)
        total_abertos += len(abertos)
        if fixos and not pacientes:
            total_alunos_sem_paciente += 1

        alunos.append({
            'id': aluno['id'],
            'username': aluno['username'],
            'nome': aluno.get('nome_completo') or aluno['username'],
            'email': aluno.get('email') or '',
            'fixos': fixos,
            'pacientes': pacientes,
            'triagens': triagens,
            'abertos': abertos,
            'semana': semana,
            'total_fixos': len(fixos),
            'total_pacientes': len(pacientes),
            'total_triagens': len(triagens),
            'total_abertos': len(abertos),
            'solicitacoes': solicitacoes_por_aluno.get(aluno['id'], []),
            'ultima_solicitacao': solicitacoes_por_aluno.get(aluno['id'], [None])[0]
        })

    solicitacoes_vagas = []
    for row in solicitacoes_rows:
        solicitacoes_vagas.append(preparar_solicitacao_vaga(row))

    return render_template(
        'minha_supervisao.html',
        usuario=current_user.username,
        papel=current_user.role,
        papel_label=PAPEIS_LABEL.get(current_user.role, current_user.role),
        alunos=alunos,
        total_alunos=len(alunos),
        total_fixos=total_fixos,
        total_pacientes=total_pacientes,
        total_triagens=total_triagens,
        total_abertos=total_abertos,
        total_alunos_sem_paciente=total_alunos_sem_paciente,
        solicitacoes_vagas=solicitacoes_vagas,
        total_solicitacoes_pendentes=sum(1 for s in solicitacoes_vagas if s['status'] in ('pendente', 'em_analise', 'atendida_parcial'))
    )


@app.route('/relatorio-semanal')
@login_required
@requer_papel_page('coordenador', 'somente_leitura')
def relatorio_semanal():
    inicio_dt, fim_dt = relatorios.periodo_semana_atual()
    inicio = inicio_dt.strftime('%Y-%m-%d')
    fim = fim_dt.strftime('%Y-%m-%d')
    supervisor_id = normalizar_supervisor_id(request.args.get('supervisor_id'))
    if supervisor_id == 'invalido':
        flash('Supervisor inválido.', 'error')
        return redirect(url_for('relatorio_semanal'))

    conn = get_db()
    try:
        dados = relatorios.coletar_relatorio_semanal(
            conn,
            inicio,
            fim,
            supervisor_id,
            listar_professores_ativos,
            DIAS,
            DIAS_PT,
            label_status_reserva,
        )
    finally:
        conn.close()

    return render_template(
        'relatorio_semanal.html',
        usuario=current_user.username,
        papel=current_user.role,
        papel_label=PAPEIS_LABEL.get(current_user.role, current_user.role),
        periodo=f'{inicio_dt.strftime("%d/%m/%Y")} a {fim_dt.strftime("%d/%m/%Y")}',
        total_ocupados=dados['total_ocupados'],
        pontuais=dados['pontuais'],
        abertos=dados['abertos'],
        reservas_sala=dados['reservas_sala'],
        reservas_instrumento=dados['reservas_instrumento'],
        dias_relatorio=dados['dias_relatorio'],
        professores=dados['professores'],
        supervisor_id=supervisor_id,
        por_supervisor=dados['por_supervisor'],
        versao=VERSAO
    )


@app.route('/meus-agendamentos')
@login_required
@requer_papel_page('aluno')
def meus_agendamentos():
    conn = get_db()
    try:
        rows = conn.execute(
            """
            SELECT *
            FROM agendamentos
            WHERE (usuario_id = ? OR (usuario_id IS NULL AND estagiario = ?))
              AND (
                data_especifica IS NULL
                OR data_especifica = ''
                OR data_especifica >= ?
              )
            ORDER BY
              CASE dia_semana
                WHEN 'SEGUNDA' THEN 1
                WHEN 'TERÇA' THEN 2
                WHEN 'QUARTA' THEN 3
                WHEN 'QUINTA' THEN 4
                WHEN 'SEXTA' THEN 5
                ELSE 6
              END,
              horario,
              sala,
              data_especifica
            """,
            (current_user.id, current_user.username, data_hoje_iso())
        ).fetchall()
        reservas_pendentes_aluno = conn.execute(
            """
            SELECT COUNT(*) AS total
            FROM reservas
            WHERE usuario_id=? AND status='pendente'
            """
            ,
            (current_user.id,)
        ).fetchone()['total']
        reservas_aluno = conn.execute(
            """
            SELECT *
            FROM reservas
            WHERE usuario_id=?
              AND status!='recusada'
              AND data_uso>=?
            ORDER BY data_uso, horario_inicio
            """,
            (current_user.id, data_hoje_iso())
        ).fetchall()
    finally:
        conn.close()

    agendamentos = []
    atendimentos_paciente = []
    horarios_fixos = []
    pacientes_pontuais_por_horario = {}

    for row in rows:
        ag = dict(row)
        ag['dia_visual'] = ag['dia_semana']
        if ag.get('data_especifica'):
            try:
                ag['dia_visual'] = dia_semana_da_data(ag['data_especifica'])
            except ValueError:
                ag['dia_visual'] = ag['dia_semana']
        ag['dia_label'] = DIAS_PT.get(ag['dia_visual'], ag['dia_visual'].title())
        ag['tipo'] = 'Pontual' if ag.get('data_especifica') else 'Fixo semanal'
        ag['tipo_slug'] = 'pontual' if ag.get('data_especifica') else 'fixo'
        ag['tem_paciente'] = bool((ag.get('paciente') or '').strip())
        ag['eh_fixo'] = not bool(ag.get('data_especifica'))
        ag['eh_triagem'] = bool(int(ag.get('triagem') or 0))
        ag['data_label'] = ''
        if ag.get('data_especifica'):
            try:
                ag['data_label'] = datetime.strptime(ag['data_especifica'], '%Y-%m-%d').strftime('%d/%m/%Y')
            except ValueError:
                ag['data_label'] = ag['data_especifica']
        ag['paciente_label_aluno'] = 'Triagem marcada' if ag['eh_triagem'] else 'Paciente marcado'
        ag['observacao_aluno'] = ''

        if ag['tem_paciente']:
            atendimentos_paciente.append(ag)
            if ag.get('data_especifica'):
                chave = (ag['dia_semana'], ag['horario'], ag['sala'])
                pacientes_pontuais_por_horario.setdefault(chave, []).append(ag)
        agendamentos.append(ag)

    for ag in agendamentos:
        if not ag['eh_fixo']:
            continue

        chave = (ag['dia_semana'], ag['horario'], ag['sala'])
        pacientes_pontuais = pacientes_pontuais_por_horario.get(chave, [])
        ag['paciente_pontual_label'] = ''
        if pacientes_pontuais:
            primeiro = pacientes_pontuais[0]
            ag['paciente_pontual_label'] = (
                f"Uso pontual em {primeiro.get('data_label') or primeiro.get('data_especifica')}"
            )

        if ag['tem_paciente']:
            continue

        categoria_upper = (ag.get('categoria') or '').upper()
        if pacientes_pontuais:
            ag['status_fixo'] = 'Fechado em data pontual'
            ag['status_slug'] = 'com-paciente'
            ag['status_descricao'] = 'Este horário fixo recebeu um paciente em uma data específica.'
        elif ag['eh_triagem']:
            ag['status_fixo'] = 'Triagem sem paciente marcado'
            ag['status_slug'] = 'aguardando'
            ag['status_descricao'] = 'Horário reservado para triagem, mas ainda sem paciente vinculado.'
        elif categoria_upper == 'MARCAR':
            ag['status_fixo'] = 'Horário aberto sem paciente'
            ag['status_slug'] = 'aberto'
            ag['status_descricao'] = 'Horário liberado para paciente, mas a recepção ainda não marcou ninguém.'
        else:
            ag['status_fixo'] = 'Reservado sem paciente'
            ag['status_slug'] = 'livre'
            ag['status_descricao'] = 'Horário fixo reservado, mas sem paciente marcado.'

        horarios_fixos.append(ag)

    status_reserva_labels = {
        'pendente': 'Aguardando análise',
        'aprovada': 'Aprovada',
        'recusada': 'Recusada',
        'separado': 'Instrumento separado',
        'guardado': 'Instrumento guardado',
        'retirado': 'Instrumento retirado',
        'devolvido': 'Instrumento devolvido',
    }
    reservas_aluno_formatadas = []
    for row in reservas_aluno:
        reserva = dict(row)
        reserva['status_label'] = status_reserva_labels.get(reserva.get('status'), reserva.get('status') or '')
        reserva['tipo_label'] = 'Sala' if reserva.get('tipo') == 'sala' else 'Instrumento'
        reserva['detalhe'] = (
            reserva.get('sala_atribuida')
            or reserva.get('sala_sugerida')
            or reserva.get('instrumento')
            or reserva.get('tipo_sala')
            or '-'
        )
        try:
            reserva['data_label'] = datetime.strptime(reserva['data_uso'], '%Y-%m-%d').strftime('%d/%m/%Y')
        except (ValueError, TypeError):
            reserva['data_label'] = reserva.get('data_uso') or ''
        reservas_aluno_formatadas.append(reserva)

    semana_aluno = []
    for dia in DIAS:
        itens = []
        for ag in atendimentos_paciente:
            if ag.get('dia_visual') != dia:
                continue
            item = dict(ag)
            item['semana_tipo'] = 'paciente'
            item['semana_titulo'] = ag.get('paciente_label_aluno') or 'Paciente marcado'
            item['semana_descricao'] = 'Triagem marcada' if ag.get('eh_triagem') else 'Atendimento com paciente'
            itens.append(item)
        for ag in horarios_fixos:
            if ag.get('dia_visual') != dia:
                continue
            item = dict(ag)
            item['semana_tipo'] = 'fixo'
            if ag.get('eh_triagem'):
                item['semana_titulo'] = 'Triagem livre'
            elif (ag.get('categoria') or '').upper() == 'MARCAR':
                item['semana_titulo'] = 'Horário aberto'
            else:
                item['semana_titulo'] = 'Horário reservado'
            item['semana_descricao'] = ag.get('status_descricao') or 'Horário reservado, ainda sem paciente.'
            itens.append(item)

        itens.sort(key=lambda item: (HORARIOS.index(item['horario']) if item.get('horario') in HORARIOS else 99, item.get('sala') or ''))
        semana_aluno.append({
            'dia': dia,
            'label': DIAS_PT.get(dia, dia.title()),
            'itens': itens
        })

    return render_template(
        'meus_agendamentos.html',
        agendamentos=agendamentos,
        atendimentos_paciente=atendimentos_paciente,
        horarios_fixos=horarios_fixos,
        semana_aluno=semana_aluno,
        total=len(agendamentos),
        total_fixos=len(horarios_fixos),
        total_atendimentos=len(atendimentos_paciente),
        total_pontuais=sum(1 for ag in atendimentos_paciente if ag.get('data_especifica')),
        reservas_aluno=reservas_aluno_formatadas,
        total_reservas_aluno=len(reservas_aluno_formatadas),
        total_triagens_livres=sum(
            1 for ag in horarios_fixos
            if ag.get('eh_triagem') and not ag.get('tem_paciente') and not ag.get('paciente_pontual_label')
        ),
        reservas_pendentes_aluno=reservas_pendentes_aluno,
        usuario=current_user.username,
        papel=current_user.role,
        papel_label=PAPEIS_LABEL.get(current_user.role, current_user.role)
    )


reservas_mod.registrar_rotas_reservas(app, {
    'get_db': get_db,
    'login_required': login_required,
    'requer_papel': requer_papel,
    'data_hoje_iso': data_hoje_iso,
    'normalizar_data_especifica': normalizar_data_especifica,
    'dia_semana_da_data': dia_semana_da_data,
    'checar_conflito': checar_conflito,
    'inserir_agendamento': inserir_agendamento,
    'detect_sem': detect_sem,
    'registrar_log': registrar_log,
    'notificar_reserva_email': notificar_reserva_email,
    'notificar_reserva_solicitada_email': notificar_reserva_solicitada_email,
    'HORARIOS': HORARIOS,
    'SALAS': SALAS,
    'SALAS_RESERVAVEIS': SALAS_RESERVAVEIS,
    'SALAS_COM_COMPUTADOR': SALAS_COM_COMPUTADOR,
    'PAPEIS_LABEL': PAPEIS_LABEL,
})

@app.route('/perfil', methods=['GET', 'POST'])
@login_required
def perfil():
    conn = get_db()
    try:
        row = conn.execute('SELECT * FROM usuarios WHERE id=?', (current_user.id,)).fetchone()
    finally:
        conn.close()

    if request.method == 'POST':
        nome_completo = request.form.get('nome_completo', '').strip()
        email = request.form.get('email', '').strip()
        if not nome_completo:
            flash('Nome completo é obrigatório.', 'error')
            return redirect(url_for('perfil'))
        if not email:
            flash('E-mail é obrigatório.', 'error')
            return redirect(url_for('perfil'))
        conn = get_db()
        try:
            conn.execute(
                'UPDATE usuarios SET nome_completo=?, email=? WHERE id=?',
                (nome_completo, email, current_user.id)
            )
            conn.commit()
        finally:
            conn.close()
        registrar_log('EDITAR_PERFIL', f'Usuário {current_user.username} atualizou o perfil')
        flash('Perfil atualizado com sucesso!', 'success')
        return redirect(url_for('perfil'))

    return render_template(
        'perfil.html',
        usuario=current_user.username,
        papel=current_user.role,
        papel_label=PAPEIS_LABEL.get(current_user.role, current_user.role),
        nome_completo=row['nome_completo'] or '',
        email=row['email'] or ''
    )


@app.route('/trocar-senha', methods=['GET', 'POST'])
@login_required
def trocar_senha():
    if request.method == 'POST':
        senha_atual = request.form.get('senha_atual', '')
        nova_senha = request.form.get('nova_senha', '').strip()
        confirmar = request.form.get('confirmar_senha', '').strip()
        conn = get_db()
        try:
            row = conn.execute('SELECT * FROM usuarios WHERE id=?', (current_user.id,)).fetchone()
        finally:
            conn.close()
        if not check_password_hash(row['password_hash'], senha_atual):
            flash('Senha atual incorreta.', 'error')
            return redirect(url_for('trocar_senha'))
        if len(nova_senha) < 8:
            flash('A nova senha deve ter no mínimo 8 caracteres.', 'error')
            return redirect(url_for('trocar_senha'))
        if nova_senha != confirmar:
            flash('As senhas não coincidem.', 'error')
            return redirect(url_for('trocar_senha'))
        conn = get_db()
        try:
            conn.execute(
                'UPDATE usuarios SET password_hash=? WHERE id=?',
                (generate_password_hash(nova_senha), current_user.id)
            )
            conn.commit()
        finally:
            conn.close()
        registrar_log('TROCAR_SENHA', f'Usuário {current_user.username} alterou a própria senha')
        try:
            if not notificar_senha_alterada_email(row, 'troca'):
                registrar_log(
                    'EMAIL_NAO_ENVIADO',
                    f'Aviso de senha alterada não enviado para {current_user.username}: SMTP indisponível ou e-mail ausente'
                )
        except Exception as exc:
            registrar_log(
                'EMAIL_ERRO',
                f'Falha ao enviar aviso de senha alterada para {current_user.username}: {exc}'
            )
        flash('Senha alterada com sucesso!', 'success')
        return redirect(url_for('perfil'))
    return render_template('trocar_senha.html', usuario=current_user.username, papel=current_user.role)


@app.route('/imprimir')
@login_required
@requer_papel_page('coordenador', 'recepcao', 'somente_leitura')
def imprimir_selecao():
    return render_template(
        'imprimir_selecao.html',
        dias=DIAS,
        dias_pt=DIAS_PT,
        usuario=current_user.username,
        papel=current_user.role
    )


@app.route('/imprimir/<dia>')
@login_required
@requer_papel_page('coordenador', 'recepcao', 'somente_leitura')
def imprimir(dia):
    dia = dia.upper()
    if dia not in DIAS:
        return 'Dia inválido', 400
    conn = get_db()
    try:
        rows = conn.execute(
            "SELECT horario, paciente FROM agendamentos "
            "WHERE dia_semana=? AND TRIM(paciente) != '' "
            "ORDER BY horario, paciente COLLATE NOCASE",
            (dia,)
        ).fetchall()
    finally:
        conn.close()
    pacientes = [{'horario': r['horario'], 'paciente': r['paciente'].strip().title()} for r in rows]
    return render_template(
        'imprimir.html',
        dia_nome=DIAS_PT.get(dia, dia),
        gerado_em=datetime.now().strftime('%d/%m/%Y %H:%M'),
        pacientes=pacientes
    )


@app.route('/logs')
@login_required
@requer_papel_page('coordenador', 'somente_leitura')
def logs_page():
    return render_template('logs.html', usuario=current_user.username, papel=current_user.role)


@app.route('/saude')
@login_required
@requer_papel_page('coordenador', 'somente_leitura')
def saude_sistema():
    return render_template(
        'saude.html',
        usuario=current_user.username,
        papel=current_user.role,
        saude=coletar_saude_sistema()
    )


# ========================================
# ROTAS DE USUARIOS
# ========================================

@app.route('/usuarios')
@login_required
@requer_papel_page('coordenador', 'somente_leitura')
def usuarios_page():
    conn = get_db()
    try:
        rows = selecionar_usuarios_para_admin(conn)
        professores = listar_professores_ativos(conn)
    finally:
        conn.close()
    return render_template(
        'usuarios.html',
        usuarios=[dict(r) for r in rows],
        professores=[dict(r) for r in professores],
        usuario=current_user.username,
        papel=current_user.role,
        papeis_label=PAPEIS_LABEL
    )


@app.route('/api/estagiarios', methods=['GET'])
@login_required
@requer_papel('coordenador', 'recepcao', 'somente_leitura')
def api_list_estagiarios():
    conn = get_db()
    try:
        rows = conn.execute(
            """
            SELECT u.id, u.username, u.supervisor_id,
                   COALESCE(NULLIF(p.nome_completo, ''), p.username, '') AS supervisor_nome
            FROM usuarios u
            LEFT JOIN usuarios p ON p.id = u.supervisor_id
            WHERE u.role='aluno' AND u.ativo=1
            ORDER BY u.username
            """
        ).fetchall()
    finally:
        conn.close()
    return jsonify([agendamento_para_resposta(r) for r in rows])


@app.route('/api/usuarios', methods=['GET'])
@login_required
@requer_papel('coordenador', 'somente_leitura')
def api_list_usuarios():
    conn = get_db()
    try:
        rows = selecionar_usuarios_para_admin(conn)
    finally:
        conn.close()
    return jsonify([dict(r) for r in rows])


@app.route('/api/usuarios', methods=['POST'])
@login_required
@requer_papel('coordenador')
@limiter.limit('20 per minute')
def api_criar_usuario():
    d = request.get_json(silent=True)
    if not d:
        return jsonify({'erro': 'Não consegui ler os dados enviados. Recarregue a página e tente novamente.'}), 400

    nome_completo = (d.get('nome_completo') or '').strip()
    username = sugerir_username_por_nome(nome_completo) or (d.get('username') or '').strip()
    email = (d.get('email') or '').strip()
    password = (d.get('password') or '').strip()
    enviar_convite = bool(d.get('enviar_convite'))
    role = (d.get('role') or 'aluno').strip()
    ativo = valor_ativo(d.get('ativo'), 1)
    supervisor_id = normalizar_supervisor_id(d.get('supervisor_id'))
    if supervisor_id == 'invalido':
        return jsonify({'erro': 'Supervisor inválido'}), 400

    if not nome_completo:
        return jsonify({'erro': 'Nome completo é obrigatório'}), 400
    if not username:
        return jsonify({'erro': 'Usuário é obrigatório'}), 400
    erro_email = validar_email_usuario(email, obrigatorio=enviar_convite)
    if erro_email:
        return jsonify({'erro': erro_email}), 400
    if not password and not (enviar_convite and email):
        return jsonify({'erro': 'Informe uma senha ou marque para enviar convite por e-mail'}), 400
    if password and len(password) < 8:
        return jsonify({'erro': 'A senha deve ter no mínimo 8 caracteres'}), 400
    if role not in PAPEIS_VALIDOS:
        return jsonify({'erro': 'Papel inválido'}), 400
    if role != 'aluno':
        supervisor_id = None

    try:
        conn = get_db()
        email_convite = None
        email_conta_criada = None
        try:
            senha_inicial = password or secrets.token_urlsafe(18)
            if supervisor_id:
                professor = conn.execute(
                    "SELECT id FROM usuarios WHERE id=? AND role='professor' AND ativo=1",
                    (supervisor_id,)
                ).fetchone()
                if not professor:
                    return jsonify({'erro': 'Supervisor professor não encontrado'}), 400
            conn.execute(
                'INSERT INTO usuarios(username, nome_completo, email, password_hash, role, supervisor_id, ativo) VALUES(?,?,?,?,?,?,?)',
                (username, nome_completo, email, generate_password_hash(senha_inicial), role, supervisor_id, ativo)
            )
            usuario_id = conn.execute('SELECT last_insert_rowid()').fetchone()[0]
            email_enviado = False
            if enviar_convite and email:
                usuario_row = conn.execute('SELECT * FROM usuarios WHERE id=?', (usuario_id,)).fetchone()
                email_convite = preparar_convite_criacao_conta(conn, usuario_row)
            elif email:
                usuario_row = conn.execute('SELECT * FROM usuarios WHERE id=?', (usuario_id,)).fetchone()
                email_conta_criada = preparar_aviso_conta_criada(usuario_row)
            conn.commit()
        finally:
            conn.close()
        if email_convite:
            email_enviado = enviar_email(*email_convite)
        elif email_conta_criada:
            email_enviado = enviar_email(*email_conta_criada)
        registrar_log('CRIAR_USUARIO', f'Usuário "{username}" ({role}) criado')
        return jsonify({'message': 'Usuário criado', 'email_enviado': email_enviado}), 201
    except sqlite3.IntegrityError:
        return jsonify({'erro': 'Nome de usuário já existe'}), 409


@app.route('/api/usuarios/<int:uid>', methods=['PUT'])
@login_required
@requer_papel('coordenador')
@limiter.limit('20 per minute')
def api_editar_usuario(uid):
    d = request.get_json(silent=True)
    if not d:
        return jsonify({'erro': 'Não consegui ler os dados enviados. Recarregue a página e tente novamente.'}), 400

    conn = get_db()
    try:
        row = conn.execute('SELECT * FROM usuarios WHERE id=?', (uid,)).fetchone()
        if not row:
            return jsonify({'erro': 'Usuário não encontrado'}), 404
        new_nome_completo = (d.get('nome_completo', row['nome_completo']) or '').strip()
        new_username = sugerir_username_por_nome(new_nome_completo) or (d.get('username', row['username']) or '').strip()
        new_email = (d.get('email', row['email']) or '').strip()
        if not new_nome_completo:
            return jsonify({'erro': 'Nome completo é obrigatório'}), 400
        if not new_username:
            return jsonify({'erro': 'Informe o nome de usuário.'}), 400
        erro_email = validar_email_usuario(new_email, obrigatorio=bool(d.get('enviar_convite')))
        if erro_email:
            return jsonify({'erro': erro_email}), 400
        new_role = (d.get('role') or row['role']).strip()
        if new_role not in PAPEIS_VALIDOS:
            return jsonify({'erro': 'Papel inválido'}), 400
        ativo = valor_ativo(d.get('ativo'), row['ativo'])
        supervisor_id = normalizar_supervisor_id(d.get('supervisor_id', row['supervisor_id']))
        if supervisor_id == 'invalido':
            return jsonify({'erro': 'Supervisor inválido'}), 400
        if new_role != 'aluno':
            supervisor_id = None
        if supervisor_id:
            professor = conn.execute(
                "SELECT id FROM usuarios WHERE id=? AND role='professor' AND ativo=1",
                (supervisor_id,)
            ).fetchone()
            if not professor:
                return jsonify({'erro': 'Supervisor professor não encontrado'}), 400
        if row['id'] == current_user.id and not ativo:
            return jsonify({'erro': 'Você não pode inativar sua própria conta'}), 400
        new_pass = (d.get('password') or '').strip()
        enviar_convite = bool(d.get('enviar_convite'))
        if enviar_convite and not new_email:
            return jsonify({'erro': 'Para enviar convite, informe o e-mail do usuário.'}), 400
        if new_pass and len(new_pass) < 8:
            return jsonify({'erro': 'A senha deve ter no mínimo 8 caracteres'}), 400
        if new_pass:
            conn.execute(
                'UPDATE usuarios SET username=?, nome_completo=?, email=?, role=?, supervisor_id=?, ativo=?, password_hash=? WHERE id=?',
                (new_username, new_nome_completo, new_email, new_role, supervisor_id, ativo, generate_password_hash(new_pass), uid)
            )
        else:
            conn.execute(
                'UPDATE usuarios SET username=?, nome_completo=?, email=?, role=?, supervisor_id=?, ativo=? WHERE id=?',
                (new_username, new_nome_completo, new_email, new_role, supervisor_id, ativo, uid)
            )
        email_convite = None
        if enviar_convite:
            usuario_row = conn.execute('SELECT * FROM usuarios WHERE id=?', (uid,)).fetchone()
            email_convite = preparar_convite_criacao_conta(conn, usuario_row)
        try:
            conn.commit()
        except Exception:
            return jsonify({'erro': 'Nome de usuário já existe'}), 400
        email_enviado = enviar_email(*email_convite) if email_convite else False
        registrar_log('EDITAR_USUARIO', f'Usuário "{row["username"]}" atualizado')
        return jsonify({'message': 'Usuário atualizado', 'email_enviado': email_enviado})
    finally:
        conn.close()


@app.route('/api/usuarios/<int:uid>', methods=['DELETE'])
@login_required
@requer_papel('coordenador')
@limiter.limit('10 per minute')
def api_excluir_usuario(uid):
    conn = get_db()
    try:
        row = conn.execute('SELECT * FROM usuarios WHERE id=?', (uid,)).fetchone()
        if not row:
            return jsonify({'erro': 'Usuário não encontrado'}), 404
        if row['id'] == current_user.id:
            return jsonify({'erro': 'Você não pode excluir sua própria conta'}), 400
        conn.execute('UPDATE usuarios SET ativo=0 WHERE id=?', (uid,))
        conn.commit()
    finally:
        conn.close()
    registrar_log('INATIVAR_USUARIO', f'Usuário "{row["username"]}" inativado')
    return jsonify({'message': 'Usuário inativado'})


@app.route('/api/usuarios/<int:uid>/excluir-definitivo', methods=['DELETE'])
@login_required
@requer_papel('coordenador')
@limiter.limit('10 per minute')
def api_excluir_usuario_definitivo(uid):
    if uid == current_user.id:
        return jsonify({'erro': 'Você não pode excluir sua própria conta'}), 400

    conn = get_db()
    try:
        row = conn.execute('SELECT * FROM usuarios WHERE id=?', (uid,)).fetchone()
        if not row:
            return jsonify({'erro': 'Usuário não encontrado'}), 404

        usos = {
            'agendamentos': conn.execute(
                'SELECT COUNT(*) AS total FROM agendamentos WHERE usuario_id=?',
                (uid,)
            ).fetchone()['total'],
            'reservas': conn.execute(
                'SELECT COUNT(*) AS total FROM reservas WHERE usuario_id=?',
                (uid,)
            ).fetchone()['total'],
        }
        if any(usos.values()):
            return jsonify({
                'erro': 'Este usuário já possui agendamentos ou reservas. Para manter o histórico correto, inative em vez de excluir.'
            }), 400

        conn.execute('UPDATE usuarios SET supervisor_id=NULL WHERE supervisor_id=?', (uid,))
        conn.execute('DELETE FROM usuarios WHERE id=?', (uid,))
        conn.commit()
    finally:
        conn.close()

    registrar_log('EXCLUIR_USUARIO', f'Usuário "{row["username"]}" excluído definitivamente')
    return jsonify({'message': 'Usuário excluído definitivamente'})


# ========================================
# API DE AGENDAMENTOS
# ========================================

@app.route('/api/conflito', methods=['GET'])
@login_required
def api_conflito():
    dia = request.args.get('dia_semana', '')
    horario = request.args.get('horario', '')
    sala = request.args.get('sala', '')
    data_esp = request.args.get('data_especifica', '').strip()
    excluir = request.args.get('excluir_id', None)
    ocupa_manual = valor_ocupa_sala(request.args.get('ocupa_sala'), None)
    if ocupa_manual is None:
        ocupa_manual = calcular_ocupa_sala(
            request.args.get('categoria', ''),
            request.args.get('paciente', ''),
            request.args.get('observacao', ''),
            data_esp,
            request.args.get('triagem', 0)
        )
    if not ocupa_manual:
        return jsonify({'conflito': False, 'ocupa_sala': False})
    if not dia or not horario or not sala:
        return jsonify({'conflito': False})
    erro_validacao = validar_valores_agendamento(dia, horario, sala)
    if erro_validacao:
        return jsonify({'erro': erro_validacao}), 400
    if data_esp:
        data_esp, erro_data = normalizar_data_especifica(data_esp)
        if erro_data:
            return jsonify({'erro': erro_data}), 400
        dia = dia_semana_da_data(data_esp)
    conflito = checar_conflito(dia, horario, sala, data_especifica=data_esp, excluir_id=excluir)
    if conflito:
        if current_user.role in ('aluno', 'professor'):
            return jsonify({'conflito': True})
        return jsonify({
            'conflito': True,
            'estagiario': conflito.get('estagiario', ''),
            'paciente': conflito.get('paciente', ''),
            'categoria': conflito.get('categoria', ''),
            'id': conflito.get('id')
        })
    return jsonify({'conflito': False})


@app.route('/api/agendamentos', methods=['GET'])
@login_required
def list_ag():
    dia = request.args.get('dia_semana', 'SEGUNDA')
    horario = request.args.get('horario', '')
    sala = request.args.get('sala', '')
    cat = request.args.get('categoria', '')
    ocupa_sala = request.args.get('ocupa_sala', '').strip()
    busca = request.args.get('busca', '').strip()
    data_ref = request.args.get('data', '').strip()
    supervisor_id = normalizar_supervisor_id(request.args.get('supervisor_id'))
    if supervisor_id == 'invalido':
        return jsonify({'erro': 'Supervisor inválido.'}), 400

    dia_ref = None
    if data_ref:
        data_ref, erro_data = normalizar_data_especifica(data_ref)
        if erro_data:
            return jsonify({'erro': erro_data}), 400
        dia_ref = dia_semana_da_data(data_ref)

    dia_busca = dia_ref or dia
    erro_validacao = validar_valores_agendamento(
        dia_busca,
        horario or HORARIOS[0],
        sala or SALAS[0],
        cat
    )
    if erro_validacao:
        return jsonify({'erro': erro_validacao}), 400

    if data_ref:
        q = ('SELECT a.*, COALESCE(aluno.supervisor_id, aluno_por_nome.supervisor_id) AS supervisor_id, '
             "COALESCE(NULLIF(prof.nome_completo, ''), prof.username, '') AS supervisor_nome "
             'FROM agendamentos a '
             'LEFT JOIN usuarios aluno ON aluno.id = a.usuario_id '
             "LEFT JOIN usuarios aluno_por_nome ON aluno_por_nome.username = a.estagiario AND aluno_por_nome.role='aluno' AND a.usuario_id IS NULL "
             'LEFT JOIN usuarios prof ON prof.id = COALESCE(aluno.supervisor_id, aluno_por_nome.supervisor_id) '
             'WHERE ('
             '(a.dia_semana=? AND (a.data_especifica IS NULL OR a.data_especifica = \'\'))'
             ' OR a.data_especifica=?'
             ')')
        p = [dia_busca, data_ref]
    else:
        q = (
            'SELECT a.*, COALESCE(aluno.supervisor_id, aluno_por_nome.supervisor_id) AS supervisor_id, '
            "COALESCE(NULLIF(prof.nome_completo, ''), prof.username, '') AS supervisor_nome "
            'FROM agendamentos a '
            'LEFT JOIN usuarios aluno ON aluno.id = a.usuario_id '
            "LEFT JOIN usuarios aluno_por_nome ON aluno_por_nome.username = a.estagiario AND aluno_por_nome.role='aluno' AND a.usuario_id IS NULL "
            'LEFT JOIN usuarios prof ON prof.id = COALESCE(aluno.supervisor_id, aluno_por_nome.supervisor_id) '
            'WHERE ('
            '(a.dia_semana=? AND (a.data_especifica IS NULL OR a.data_especifica = \'\')) '
            'OR (a.data_especifica IS NOT NULL AND a.data_especifica != \'\' '
            'AND a.data_especifica >= ? AND strftime(\'%w\', a.data_especifica)=?)'
            ')'
        )
        p = [dia_busca, data_hoje_iso(), numero_semana_sqlite(dia_busca)]
    if horario:
        q += ' AND a.horario=?'
        p.append(horario)
    if sala:
        q += ' AND a.sala=?'
        p.append(sala)
    if cat:
        q += ' AND a.categoria=?'
        p.append(cat)
    if ocupa_sala in ('0', '1'):
        q += ' AND a.ocupa_sala=?'
        p.append(int(ocupa_sala))
    if supervisor_id:
        q += ' AND COALESCE(aluno.supervisor_id, aluno_por_nome.supervisor_id)=?'
        p.append(supervisor_id)
    if busca:
        if current_user.role == 'aluno':
            q += ' AND (a.estagiario LIKE ? OR a.categoria LIKE ? OR a.sala LIKE ?)'
            p += [f'%{busca}%'] * 3
        else:
            q += ' AND (a.estagiario LIKE ? OR a.paciente LIKE ? OR a.observacao LIKE ? OR prof.nome_completo LIKE ? OR prof.username LIKE ?)'
            p += [f'%{busca}%'] * 3
            p += [f'%{busca}%'] * 2
    if current_user.role == 'aluno':
        q += ' AND (a.usuario_id = ? OR (a.usuario_id IS NULL AND a.estagiario = ?))'
        p += [current_user.id, current_user.username]
    elif current_user.role == 'professor':
        q += ' AND COALESCE(aluno.supervisor_id, aluno_por_nome.supervisor_id)=?'
        p.append(current_user.id)
    q += (
        ' ORDER BY a.horario, a.sala, '
        'a.ocupa_sala DESC, '
        "CASE WHEN a.data_especifica IS NOT NULL AND a.data_especifica != '' THEN 0 ELSE 1 END, "
        'a.data_especifica, a.id'
    )
    conn = get_db()
    try:
        rows = conn.execute(q, p).fetchall()
    finally:
        conn.close()
    return jsonify([agendamento_para_resposta(r) for r in rows])


@app.route('/api/agendamentos/<int:aid>', methods=['GET'])
@login_required
def get_ag(aid):
    conn = get_db()
    try:
        r = conn.execute(
            """
            SELECT a.*, COALESCE(aluno.supervisor_id, aluno_por_nome.supervisor_id) AS supervisor_id,
                   COALESCE(NULLIF(prof.nome_completo, ''), prof.username, '') AS supervisor_nome
            FROM agendamentos a
            LEFT JOIN usuarios aluno ON aluno.id = a.usuario_id
            LEFT JOIN usuarios aluno_por_nome ON aluno_por_nome.username = a.estagiario
                 AND aluno_por_nome.role='aluno' AND a.usuario_id IS NULL
            LEFT JOIN usuarios prof ON prof.id = COALESCE(aluno.supervisor_id, aluno_por_nome.supervisor_id)
            WHERE a.id=?
            """,
            (aid,)
        ).fetchone()
    finally:
        conn.close()
    if not usuario_pode_ver_agendamento(r):
        return jsonify({'erro': 'Não encontrado'}), 404
    return jsonify(agendamento_para_resposta(r))


@app.route('/api/agendamentos', methods=['POST'])
@login_required
@requer_papel('coordenador', 'recepcao')
@limiter.limit('60 per minute')
def create_ag():
    d = request.get_json(silent=True)
    if not d:
        return jsonify({'erro': 'Não consegui ler os dados do agendamento. Recarregue a página e tente novamente.'}), 400

    dados_ag, erro = preparar_dados_agendamento(d, current_user.id if current_user.is_authenticated else None)
    if erro:
        return jsonify({'erro': erro}), 400

    conflito = checar_conflito(
        dados_ag['dia'],
        dados_ag['horario'],
        dados_ag['sala'],
        data_especifica=dados_ag['data_especifica']
    ) if dados_ag['ocupa_sala'] else None
    if conflito:
        ocu = conflito.get('estagiario') or conflito.get('categoria') or 'outro agendamento'
        return jsonify({
            'erro': (
                f'Conflito: {dados_ag["sala"]} já está ocupada às {dados_ag["horario"]} '
                f'({dados_ag["data_especifica"] or dados_ag["dia"]}) por: {ocu}'
            ),
            'conflito': True,
            'conflito_id': conflito.get('id')
        }), 409

    try:
        conn = get_db()
        try:
            nid = inserir_agendamento(conn, dados_ag)
            row_notificacao = conn.execute('SELECT * FROM agendamentos WHERE id=?', (nid,)).fetchone()
            conn.commit()
        finally:
            conn.close()
    except sqlite3.IntegrityError:
        return jsonify({'erro': 'Conflito: esse horário já foi ocupado por outro agendamento.'}), 409

    registrar_log('CRIAR', descricao_agendamento_log(nid, dados_ag) + ' criado')
    if row_notificacao:
        notificar_agendamento_email(dict(row_notificacao), 'criado')
    return jsonify({'id': nid, 'message': 'Criado'}), 201


@app.route('/api/agendamentos/<int:aid>', methods=['PUT'])
@login_required
@requer_papel('coordenador', 'recepcao')
@limiter.limit('60 per minute')
def update_ag(aid):
    d = request.get_json(silent=True)
    if not d:
        return jsonify({'erro': 'Não consegui ler os dados do agendamento. Recarregue a página e tente novamente.'}), 400

    dados_ag, erro = preparar_dados_agendamento(d)
    if erro:
        return jsonify({'erro': erro}), 400

    conflito = checar_conflito(
        dados_ag['dia'],
        dados_ag['horario'],
        dados_ag['sala'],
        data_especifica=dados_ag['data_especifica'],
        excluir_id=aid
    ) if dados_ag['ocupa_sala'] else None
    if conflito:
        ocu = conflito.get('estagiario') or conflito.get('categoria') or 'outro agendamento'
        return jsonify({
            'erro': (
                f'Conflito: {dados_ag["sala"]} já está ocupada às {dados_ag["horario"]} '
                f'({dados_ag["data_especifica"] or dados_ag["dia"]}) por: {ocu}'
            ),
            'conflito': True,
            'conflito_id': conflito.get('id')
        }), 409

    try:
        conn = get_db()
        try:
            row_anterior = conn.execute('SELECT * FROM agendamentos WHERE id=?', (aid,)).fetchone()
            if not row_anterior:
                return jsonify({'erro': 'Agendamento não encontrado. Ele pode ter sido removido por outra pessoa.'}), 404
            cur = conn.execute(
                'UPDATE agendamentos SET dia_semana=?,horario=?,sala=?,estagiario=?,paciente=?,categoria=?,semestre=?,'
                'triagem=?,observacao=?,data_especifica=?,usuario_id=?,ocupa_sala=?,status_atendimento=?,updated_at=CURRENT_TIMESTAMP WHERE id=?',
                (
                    dados_ag['dia'], dados_ag['horario'], dados_ag['sala'],
                    dados_ag['estagiario'], dados_ag['paciente'], dados_ag['categoria'],
                    dados_ag['semestre'], dados_ag['triagem'], dados_ag['observacao'],
                    dados_ag['data_especifica'],
                    buscar_usuario_id_aluno(dados_ag['estagiario'], conn) or dados_ag['usuario_id'],
                    dados_ag['ocupa_sala'],
                    dados_ag['status_atendimento'],
                    aid
                )
            )
            if cur.rowcount == 0:
                return jsonify({'erro': 'Agendamento não encontrado. Ele pode ter sido removido por outra pessoa.'}), 404
            row_notificacao = conn.execute('SELECT * FROM agendamentos WHERE id=?', (aid,)).fetchone()
            conn.commit()
        finally:
            conn.close()
    except sqlite3.IntegrityError:
        return jsonify({'erro': 'Conflito: esse horário já foi ocupado por outro agendamento.'}), 409

    registrar_log('EDITAR', descricao_agendamento_log(aid, dados_ag) + ' editado')
    if row_notificacao:
        notificar_agendamento_email(dict(row_notificacao), 'alterado', dict(row_anterior))
    return jsonify({'message': 'Atualizado'})


@app.route('/api/agendamentos/<int:aid>', methods=['DELETE'])
@login_required
@requer_papel('coordenador')
@limiter.limit('30 per minute')
def delete_ag(aid):
    conn = get_db()
    try:
        r = conn.execute('SELECT * FROM agendamentos WHERE id=?', (aid,)).fetchone()
        conn.execute('DELETE FROM agendamentos WHERE id=?', (aid,))
        conn.commit()
    finally:
        conn.close()
    if r:
        registrar_log('EXCLUIR', descricao_agendamento_log(aid, dict(r)) + ' excluido')
        notificar_agendamento_email(dict(r), 'excluido')
    return jsonify({'message': 'Removido'})


# ========================================
# API DE RELATORIOS E ADMINISTRACAO
# ========================================

@app.route('/api/stats')
@login_required
@requer_papel('coordenador', 'recepcao', 'somente_leitura')
def stats():
    dia = request.args.get('dia_semana', 'SEGUNDA')
    supervisor_id = normalizar_supervisor_id(request.args.get('supervisor_id'))
    if dia not in DIAS:
        return jsonify({'erro': 'Escolha um dia da semana válido.'}), 400
    if supervisor_id == 'invalido':
        return jsonify({'erro': 'Supervisor inválido.'}), 400
    filtro_visao = (
        "((a.dia_semana=? AND (a.data_especifica IS NULL OR a.data_especifica = '')) "
        "OR (a.data_especifica IS NOT NULL AND a.data_especifica != '' "
        "AND a.data_especifica >= ? AND strftime('%w', a.data_especifica)=?))"
    )
    joins_supervisor = (
        " LEFT JOIN usuarios aluno ON aluno.id = a.usuario_id "
        " LEFT JOIN usuarios aluno_por_nome ON aluno_por_nome.username = a.estagiario "
        "      AND aluno_por_nome.role='aluno' AND a.usuario_id IS NULL "
    )
    filtro_supervisor = ''
    params = [dia, data_hoje_iso(), numero_semana_sqlite(dia)]
    if supervisor_id:
        filtro_supervisor = " AND COALESCE(aluno.supervisor_id, aluno_por_nome.supervisor_id)=?"
        params.append(supervisor_id)
    conn = get_db()
    try:
        total = conn.execute(
            f'SELECT COUNT(*) FROM agendamentos a {joins_supervisor} WHERE {filtro_visao}{filtro_supervisor}',
            params
        ).fetchone()[0]
        livre = conn.execute(
            f"SELECT COUNT(*) FROM agendamentos a {joins_supervisor} WHERE {filtro_visao} AND a.categoria='LIVRE'{filtro_supervisor}",
            params
        ).fetchone()[0]
        por_cat = conn.execute(
            f'SELECT a.categoria, COUNT(*) as n FROM agendamentos a {joins_supervisor} WHERE {filtro_visao}{filtro_supervisor} GROUP BY a.categoria ORDER BY n DESC',
            params
        ).fetchall()
    finally:
        conn.close()
    return jsonify({'total': total, 'livre': livre, 'por_categoria': [dict(r) for r in por_cat]})


@app.route('/api/export')
@login_required
@requer_papel('coordenador')
def export_xlsx():
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        return jsonify({
            'erro': 'Exportação XLSX indisponível. Instale as dependências com: pip install -r requirements.txt'
        }), 500

    conn = get_db()
    try:
        rows = conn.execute(
            """
            SELECT a.*, COALESCE(u.nome_completo, a.estagiario) as nome_real,
                   COALESCE(NULLIF(p.nome_completo, ''), p.username, '') AS supervisor_nome
            FROM agendamentos a
            LEFT JOIN usuarios u ON a.usuario_id = u.id
            LEFT JOIN usuarios u_nome ON u_nome.username = a.estagiario AND u_nome.role='aluno' AND a.usuario_id IS NULL
            LEFT JOIN usuarios p ON p.id = COALESCE(u.supervisor_id, u_nome.supervisor_id)
            ORDER BY a.dia_semana, a.horario, a.sala
            """
        ).fetchall()
    finally:
        conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = 'Mapa de Salas'
    headers = ['ID', 'Dia', 'Horário', 'Sala', 'Estagiário (usuário)', 'Nome Completo', 'Supervisor', 'Paciente',
               'Categoria', 'Status Atendimento', 'Semestre', 'Triagem', 'Ocupa Sala', 'Data Esp.', 'Obs.']
    ws.append(headers)

    header_fill = PatternFill('solid', fgColor='1E293B')
    header_font = Font(color='FFFFFF', bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    for r in rows:
        ws.append([
            r['id'], r['dia_semana'], r['horario'], r['sala'],
            r['estagiario'], r['nome_real'], r['supervisor_nome'], r['paciente'],
            r['categoria'], STATUS_ATENDIMENTO.get(r['status_atendimento'], ''),
            r['semestre'], 'Sim' if r['triagem'] else 'Não',
            'Sim' if r['ocupa_sala'] else 'Não', r['data_especifica'], r['observacao']
        ])

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max(max_len + 2, 10), 42)

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    registrar_log('EXPORTAR', 'XLSX exportado')
    return send_file(
        out,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'mapa_salas_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
    )


@app.route('/api/logs')
@login_required
@requer_papel('coordenador', 'somente_leitura')
def get_logs():
    pagina, erro = inteiro_query('pagina', 1, minimo=1)
    if erro:
        return jsonify({'erro': erro}), 400
    por_pag, erro = inteiro_query('por_pagina', 50, minimo=10, maximo=100)
    if erro:
        return jsonify({'erro': erro}), 400
    usuario = request.args.get('usuario', '').strip()
    acao = request.args.get('acao', '').strip()
    data_ini = request.args.get('data_ini', '').strip()
    data_fim = request.args.get('data_fim', '').strip()
    offset = (pagina - 1) * por_pag

    q = 'SELECT * FROM historico WHERE 1=1'
    p = []
    if usuario:
        q += ' AND usuario LIKE ?'
        p.append(f'%{usuario}%')
    if acao:
        q += ' AND acao LIKE ?'
        p.append(f'%{acao}%')
    if data_ini:
        q += ' AND ts >= ?'
        p.append(data_ini)
    if data_fim:
        q += ' AND ts <= ?'
        p.append(data_fim + ' 23:59:59')

    conn = get_db()
    try:
        total = conn.execute(q.replace('SELECT *', 'SELECT COUNT(*)'), p).fetchone()[0]
        rows = conn.execute(q + ' ORDER BY ts DESC LIMIT ? OFFSET ?', p + [por_pag, offset]).fetchall()
    finally:
        conn.close()
    return jsonify({
        'total': total,
        'pagina': pagina,
        'por_pagina': por_pag,
        'paginas': (total + por_pag - 1) // por_pag,
        'logs': [dict(r) for r in rows]
    })


@app.route('/api/admin/limpar-logs', methods=['POST'])
@login_required
@requer_papel('coordenador')
def limpar_logs():
    conn = get_db()
    try:
        removidos = limpar_logs_antigos(conn)
        conn.commit()
    finally:
        conn.close()
    registrar_log('LIMPAR_LOGS', f'{removidos} logs antigos removidos (retenção: {LOG_RETENCAO_DIAS} dias)')
    return jsonify({'message': f'{removidos} logs removidos'})


@app.route('/api/admin/apagar-logs-recentes', methods=['POST'])
@login_required
@requer_papel('coordenador')
def apagar_logs_recentes():
    conn = get_db()
    try:
        cur = conn.execute(
            "DELETE FROM historico WHERE ts >= datetime('now', '-' || ? || ' days')",
            (LOG_RETENCAO_DIAS,)
        )
        removidos = cur.rowcount
        conn.commit()
    finally:
        conn.close()
    registrar_log('APAGAR_LOGS_RECENTES', f'{removidos} logs dos últimos {LOG_RETENCAO_DIAS} dias removidos')
    return jsonify({'message': f'{removidos} logs dos últimos {LOG_RETENCAO_DIAS} dias removidos'})


@app.route('/api/busca')
@login_required
def busca_global():
    q_busca = request.args.get('q', '').strip()
    if not q_busca or len(q_busca) < 2:
        return jsonify({'erro': 'Busca deve ter ao menos 2 caracteres'}), 400
    like = f'%{q_busca}%'
    if current_user.role == 'aluno':
        q = ('SELECT * FROM agendamentos '
             'WHERE (estagiario LIKE ? OR categoria LIKE ? OR sala LIKE ?) '
             'AND (usuario_id = ? OR (usuario_id IS NULL AND estagiario = ?))')
        p = [like, like, like, current_user.id, current_user.username]
    else:
        q = ('SELECT * FROM agendamentos '
             'WHERE (estagiario LIKE ? OR paciente LIKE ? OR observacao LIKE ?)')
        p = [like, like, like]
    if current_user.role == 'professor':
        q += (
            ' AND EXISTS (SELECT 1 FROM usuarios u '
            'WHERE u.role=? AND u.supervisor_id=? '
            'AND (agendamentos.usuario_id=u.id OR agendamentos.estagiario=u.username OR agendamentos.estagiario=u.nome_completo))'
        )
        p += ['aluno', current_user.id]
    q += ' ORDER BY dia_semana, horario, sala'
    conn = get_db()
    try:
        rows = conn.execute(q, p).fetchall()
    finally:
        conn.close()
    return jsonify([agendamento_para_resposta(r) for r in rows])


def chave_sem_acento(valor):
    return excel_import_utils.chave_sem_acento(valor)


def normalizar_sala_excel(valor):
    return excel_import_utils.normalizar_sala_excel(valor, SALAS)


def texto_celula_excel(valor):
    return excel_import_utils.texto_celula_excel(valor)


def horario_excel(valor):
    return excel_import_utils.horario_excel(valor)


def extrair_data_pontual_excel(texto, ano=None):
    return excel_import_utils.extrair_data_pontual_excel(texto, ano)


def limpar_marcadores_excel(texto):
    return excel_import_utils.limpar_marcadores_excel(texto)


def categoria_por_texto_excel(texto):
    return excel_import_utils.categoria_por_texto_excel(texto)


def semestre_por_texto_excel(texto):
    return excel_import_utils.semestre_por_texto_excel(texto)


def texto_excel_eh_marcador_operacional(texto):
    return excel_import_utils.texto_excel_eh_marcador_operacional(texto)


def montar_agendamento_excel(dia, horario, sala, texto_principal, texto_secundario, ano=None):
    return excel_import_utils.montar_agendamento_excel(
        dia, horario, sala, texto_principal, texto_secundario, preparar_dados_agendamento, ano
    )


COLUNAS_AGENDAMENTO_BACKUP = (
    'id', 'dia_semana', 'horario', 'sala', 'estagiario', 'paciente',
    'categoria', 'semestre', 'triagem', 'observacao', 'data_especifica',
    'usuario_id', 'ocupa_sala', 'status_atendimento', 'created_at', 'updated_at'
)


def salvar_backup_antes_importacao(tipo, arquivo):
    colunas = ', '.join(COLUNAS_AGENDAMENTO_BACKUP)
    with db_connection(commit=True) as conn:
        rows = conn.execute(f'SELECT {colunas} FROM agendamentos ORDER BY id').fetchall()
        dados_json = json.dumps([dict(row) for row in rows], ensure_ascii=False)
        conn.execute(
            """
            INSERT INTO backups_importacao(usuario, tipo, arquivo, total_agendamentos, dados_json)
            VALUES(?,?,?,?,?)
            """,
            (
                current_user.username if current_user.is_authenticated else 'sistema',
                tipo,
                (arquivo or '')[:180],
                len(rows),
                dados_json
            )
        )
        conn.execute(
            """
            DELETE FROM backups_importacao
            WHERE id NOT IN (
                SELECT id
                FROM backups_importacao
                ORDER BY created_at DESC, id DESC
                LIMIT 5
            )
            """
        )


def restaurar_ultimo_backup_importacao():
    placeholders = ','.join('?' for _ in COLUNAS_AGENDAMENTO_BACKUP)
    colunas = ', '.join(COLUNAS_AGENDAMENTO_BACKUP)
    with db_connection(commit=True) as conn:
        backup = conn.execute(
            """
            SELECT *
            FROM backups_importacao
            ORDER BY created_at DESC, id DESC
            LIMIT 1
            """
        ).fetchone()
        if not backup:
            return None, 'Não há importação recente para desfazer.'

        try:
            agendamentos = json.loads(backup['dados_json'])
        except json.JSONDecodeError:
            return None, 'O backup da importação está corrompido.'

        conn.execute('DELETE FROM agendamentos')
        for agendamento in agendamentos:
            conn.execute(
                f'INSERT INTO agendamentos({colunas}) VALUES({placeholders})',
                [agendamento.get(coluna) for coluna in COLUNAS_AGENDAMENTO_BACKUP]
            )
        conn.execute('DELETE FROM backups_importacao WHERE id=?', (backup['id'],))
        return backup, None


def importar_xlsx_mapa(file_storage, substituir=False):
    try:
        from openpyxl import load_workbook
    except ImportError:
        return None, {'erro': 'Importação XLSX indisponível. Instale as dependências com: pip install -r requirements.txt'}, 500

    file_storage.stream.seek(0)
    wb = load_workbook(file_storage.stream, data_only=True, read_only=True)
    abas_dias = {chave_sem_acento(dia): dia for dia in DIAS}
    inseridos = 0
    conflitos = []
    erros = []
    ignorados = 0
    ano = datetime.now().year

    conn = get_db()
    try:
        if substituir:
            conn.execute('DELETE FROM agendamentos')

        for nome_aba in wb.sheetnames:
            dia = abas_dias.get(chave_sem_acento(nome_aba))
            if not dia:
                continue
            ws = wb[nome_aba]
            salas_colunas = {}
            for col in range(2, ws.max_column + 1):
                sala = normalizar_sala_excel(ws.cell(row=1, column=col).value)
                if sala:
                    salas_colunas[col] = sala

            row = 2
            while row <= ws.max_row:
                horario = horario_excel(ws.cell(row=row, column=1).value)
                if horario not in HORARIOS:
                    row += 1
                    continue

                for col, sala in salas_colunas.items():
                    texto_principal = texto_celula_excel(ws.cell(row=row, column=col).value)
                    texto_secundario = texto_celula_excel(ws.cell(row=row + 1, column=col).value) if row + 1 <= ws.max_row else ''
                    if not texto_principal and not texto_secundario:
                        continue

                    dados_ag, erro = montar_agendamento_excel(dia, horario, sala, texto_principal, texto_secundario, ano)
                    if erro:
                        ignorados += 1
                        continue

                    conflito = None if substituir or not dados_ag['ocupa_sala'] else checar_conflito(
                        dados_ag['dia'],
                        dados_ag['horario'],
                        dados_ag['sala'],
                        data_especifica=dados_ag['data_especifica']
                    )
                    if conflito:
                        conflitos.append({
                            'aba': nome_aba,
                            'horario': dados_ag['horario'],
                            'sala': dados_ag['sala'],
                            'ocupado_por': conflito.get('estagiario') or conflito.get('categoria')
                        })
                        continue

                    try:
                        inserir_agendamento(conn, dados_ag)
                        inseridos += 1
                    except sqlite3.IntegrityError as exc:
                        conflitos.append({
                            'aba': nome_aba,
                            'horario': dados_ag['horario'],
                            'sala': dados_ag['sala'],
                            'ocupado_por': str(exc)
                        })
                    except Exception as exc:
                        erros.append(f'{nome_aba} {horario} {sala}: {exc}')

                row += 2

        conn.commit()
    finally:
        conn.close()
        wb.close()

    registrar_log('IMPORTAR_XLSX', f'{inseridos} agendamentos importados do Excel, {len(conflitos)} conflitos, {len(erros)} erros')
    return {
        'inseridos': inseridos,
        'conflitos': conflitos[:50],
        'erros': erros[:50],
        'ignorados': ignorados,
        'substituiu': bool(substituir),
        'message': f'{inseridos} agendamento(s) importado(s) do Excel.'
    }, None, 200


@app.route('/api/import', methods=['POST'])
@login_required
@requer_papel('coordenador')
@limiter.limit('10 per minute')
def import_csv():
    if 'file' not in request.files:
        return jsonify({'erro': 'Nenhum arquivo enviado'}), 400
    f = request.files['file']
    filename = (f.filename or '').lower()
    substituir = request.form.get('substituir') in ('1', 'true', 'sim', 'on')
    if filename.endswith('.xlsx'):
        salvar_backup_antes_importacao('xlsx', filename)
        resultado, erro, status = importar_xlsx_mapa(f, substituir=substituir)
        if erro:
            return jsonify(erro), status
        return jsonify(resultado), status
    if not filename.endswith('.csv'):
        return jsonify({'erro': 'Apenas arquivos .csv ou .xlsx são aceitos'}), 400

    salvar_backup_antes_importacao('csv', filename)
    stream = io.StringIO(f.read().decode('utf-8-sig'))
    reader = csv.DictReader(stream)
    inseridos = 0
    conflitos = []
    erros = []

    conn = get_db()
    try:
        for i, row in enumerate(reader, start=2):
            try:
                dia = (row.get('Dia') or row.get('dia_semana') or '').strip().upper()
                horario = (row.get('Horário') or row.get('horario') or '').strip()
                sala = (row.get('Sala') or row.get('sala') or '').strip()
                estagiario = (row.get('Estagiário (usuário)') or row.get('Estagiário') or row.get('estagiario') or '').strip()
                paciente = (row.get('Paciente') or row.get('paciente') or '').strip()
                categoria = (row.get('Categoria') or row.get('categoria') or '').strip()
                semestre = int(row.get('Semestre') or row.get('semestre') or 0)
                triagem = 1 if str(row.get('Triagem') or '').strip().lower() in ('sim', '1', 'true') else 0
                obs = (row.get('Obs.') or row.get('observacao') or '').strip()
                data_esp = (row.get('Data Esp.') or row.get('data_especifica') or '').strip()
                ocupa_sala = row.get('Ocupa Sala') or row.get('ocupa_sala') or ''
                status_atendimento = (row.get('Status Atendimento') or row.get('status_atendimento') or '').strip()

                dados_ag, erro = preparar_dados_agendamento({
                    'dia_semana': dia,
                    'horario': horario,
                    'sala': sala,
                    'estagiario': estagiario,
                    'paciente': paciente,
                    'categoria': categoria,
                    'semestre': semestre,
                    'triagem': triagem,
                    'observacao': obs,
                    'data_especifica': data_esp,
                    'ocupa_sala': ocupa_sala,
                    'status_atendimento': status_atendimento
                })
                if erro:
                    erros.append(f'Linha {i}: {erro}')
                    continue

                conflito = checar_conflito(
                    dados_ag['dia'],
                    dados_ag['horario'],
                    dados_ag['sala'],
                    data_especifica=dados_ag['data_especifica']
                )
                if conflito:
                    conflitos.append({
                        'linha': i,
                        'dia': dados_ag['dia'],
                        'horario': dados_ag['horario'],
                        'sala': dados_ag['sala'],
                        'ocupado_por': conflito.get('estagiario') or conflito.get('categoria')
                    })
                    continue

                inserir_agendamento(conn, dados_ag)
                inseridos += 1
            except Exception as e:
                erros.append(f'Linha {i}: {str(e)}')
        conn.commit()
    finally:
        conn.close()

    registrar_log('IMPORTAR_CSV', f'{inseridos} agendamentos importados, {len(conflitos)} conflitos, {len(erros)} erros')
    return jsonify({
        'inseridos': inseridos,
        'conflitos': conflitos,
        'erros': erros,
        'message': f'{inseridos} agendamento(s) importado(s) com sucesso.'
    })


@app.route('/api/import/desfazer', methods=['POST'])
@login_required
@requer_papel('coordenador')
@limiter.limit('10 per hour')
def desfazer_importacao():
    backup, erro = restaurar_ultimo_backup_importacao()
    if erro:
        return jsonify({'erro': erro}), 404

    registrar_log(
        'DESFAZER_IMPORTACAO',
        (
            f'Importação desfeita. Arquivo anterior: {backup["arquivo"]}; '
            f'{backup["total_agendamentos"]} agendamentos restaurados.'
        )
    )
    return jsonify({
        'message': 'Última importação desfeita.',
        'restaurados': backup['total_agendamentos'],
        'arquivo': backup['arquivo']
    })


@app.route('/api/backup')
@login_required
@requer_papel('coordenador')
@limiter.limit('5 per hour')
def backup_db():
    confirm = request.args.get('confirmar', '')
    if confirm != 'sim':
        return render_template(
            'backup.html',
            usuario=current_user.username,
            papel=current_user.role,
            versao=VERSAO
        )
    backup_bytes = criar_backup_sqlite_bytes()
    ip = request.headers.get('X-Forwarded-For', request.remote_addr or '').split(',')[0].strip()
    registrar_log('BACKUP', f'Backup manual baixado por coordenador #{current_user.id} - IP: {ip}')
    return send_file(
        io.BytesIO(backup_bytes),
        mimetype='application/x-sqlite3',
        as_attachment=True,
        download_name=f'backup_mapa_{datetime.now().strftime("%Y%m%d_%H%M")}.db'
    )


@app.route('/api/backup/restaurar', methods=['POST'])
@login_required
@requer_papel('coordenador')
@limiter.limit('3 per hour')
def restaurar_backup_db():
    confirmacao = (request.form.get('confirmacao') or '').strip().upper()
    if confirmacao != 'RESTAURAR':
        flash('Digite RESTAURAR para confirmar a troca do banco.', 'error')
        return redirect(url_for('backup_db'))

    arquivo = request.files.get('backup_file')
    if not arquivo or not arquivo.filename:
        flash('Selecione um arquivo de backup .db.', 'error')
        return redirect(url_for('backup_db'))

    if not arquivo.filename.lower().endswith('.db'):
        flash('Envie apenas arquivo de backup no formato .db.', 'error')
        return redirect(url_for('backup_db'))

    tmp = tempfile.NamedTemporaryFile(prefix='restore_mapa_', suffix='.db', delete=False)
    tmp_path = tmp.name
    tmp.close()
    try:
        arquivo.save(tmp_path)
        valido, erro = validar_backup_sqlite(tmp_path)
        if not valido:
            flash(erro, 'error')
            return redirect(url_for('backup_db'))

        backup_atual = salvar_backup_antes_da_restauracao()
        remover_arquivos_sqlite_auxiliares()
        shutil.copy2(tmp_path, DB_PATH)
        executar_manutencao(vacuum=True)
        registrar_log(
            'RESTAURAR_BACKUP',
            f'Backup restaurado por coordenador #{current_user.id}. Cópia anterior salva em {backup_atual}'
        )
        flash('Backup restaurado com sucesso. Faça reload do Web App e confira o mapa, usuários e reservas.', 'success')
    except Exception as exc:
        app.logger.exception('Falha ao restaurar backup')
        flash(f'Não foi possível restaurar o backup: {exc}', 'error')
    finally:
        try:
            os.remove(tmp_path)
        except OSError:
            pass

    return redirect(url_for('backup_db'))


def executar_manutencao(vacuum=False):
    conn = get_db()
    try:
        return system_utils.executar_manutencao(conn, limpar_logs_antigos, vacuum)
    finally:
        conn.close()


@app.cli.command('manutencao')
@click.option('--vacuum', is_flag=True, help='Compacta o arquivo SQLite depois da limpeza.')
def comando_manutencao(vacuum):
    resultado = executar_manutencao(vacuum=vacuum)
    click.echo('Manutencao concluida.')
    for chave, valor in resultado.items():
        click.echo(f'{chave}: {valor}')


@app.cli.command('backup-diario')
def comando_backup_diario():
    resultado = salvar_backup_automatico()
    click.echo('Backup diario concluido.')
    click.echo(f"arquivo: {resultado['arquivo']}")
    click.echo(f"tamanho: {resultado['tamanho']} bytes")
    click.echo(f"backups_antigos_removidos: {resultado['antigos_removidos']}")
    click.echo(f"retencao_dias: {resultado['retencao_dias']}")


@app.cli.command('testar-email')
@click.option('--para', default='', help='E-mail que recebera a mensagem de teste.')
def comando_testar_email(para):
    destino = (para or SMTP_USER or EMAIL_FROM or '').strip()
    status = diagnostico_smtp()
    click.echo('Diagnostico SMTP:')
    click.echo(f"configurado: {status['configurado']}")
    click.echo(f"host: {status['host']}")
    click.echo(f"porta: {status['porta']}")
    click.echo(f"tls: {status['tls']}")
    click.echo(f"usuario_configurado: {status['usuario_configurado']}")
    click.echo(f"senha_configurada: {status['senha_configurada']}")
    click.echo(f"email_saida_configurado: {status['email_saida_configurado']}")
    if status['faltando']:
        click.echo(f"faltando: {', '.join(status['faltando'])}")
        raise click.ClickException('SMTP incompleto. Confira variaveis de ambiente no WSGI/.env.')
    if not destino:
        raise click.ClickException('Informe um destino com --para email@exemplo.com.')
    enviado = enviar_email(
        destino,
        'Teste SMTP - Mapa de Sala',
        'Este é um e-mail de teste do Mapa de Sala. Se você recebeu, o SMTP está funcionando.'
    )
    if not enviado:
        raise click.ClickException('Falha ao enviar. Veja o log EMAIL_ERRO para o detalhe técnico.')
    click.echo(f'E-mail de teste enviado para {destino}.')


# ========================================
# INICIALIZACAO DA APLICACAO
# ========================================

with app.app_context():
    init_db()

if __name__ == '__main__':
    if not app.debug or os.environ.get('WERKZEUG_RUN_MAIN') == 'true':
        init_db()
    print(f'\n Versão: {VERSAO} | http://localhost:5000\n')
    app.run(debug=True, port=5000)
