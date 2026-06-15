import sqlite3, re, sys

try:
    import openpyxl
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable,'-m','pip','install','openpyxl','--user','-q'])
    import openpyxl

SALAS = ['Consultório 1','Consultório 2','Consultório 3','Consultório 4','Consultório 5',
         'Consultório 6 (Divã)','Consultório 7 (Divã)','Consultório 8','SOU / NACE',
         'Ludoterapia','Multifuncional','Sala de Grupo 1','Sala de Grupo 2','Supervisão','Coordenação']

def normalize(t):
    if not t: return ''
    if not isinstance(t, str): t = str(t)
    for o,n in [('ă','ã'),('Ă','Ã'),('ş','º'),('Ş','º')]:
        t = t.replace(o,n)
    return t.strip()

def detect_cat(est, pac):
    c = normalize(est+' '+pac).upper()
    if 'NÃO MARCAR' in c or 'NAO MARCAR' in c: return 'NÃO MARCAR'
    if 'PSICODIAG' in c: return 'PSICODIAGNÓSTICO'
    if 'SUPERVISÃO' in c or 'SUPERVISAO' in c or 'PROF.' in c: return 'SUPERVISÃO'
    if re.search(r'PROF\s+\w', c): return 'SUPERVISÃO'
    if 'NACE' in c: return 'NACE'
    if re.search(r'\bSOU\b', c): return 'SOU'
    if 'MARCAR' in c: return 'MARCAR'
    if 'NUTRIÇÃO' in c: return 'NUTRIÇÃO'
    if 'PSIQUIATRIA' in c: return 'PSIQUIATRIA'
    if 'AMBULAT' in c: return 'AMBULATÓRIO NEUROPSICOLOGIA'
    if 'PLANTÃO' in c or 'PLANTAO' in c: return 'PLANTÃO PSICOLÓGICO'
    if 'PRONTUÁRIO' in c or 'PRONTUARIO' in c or 'ESTUDAR' in c: return 'PRONTUÁRIO/ESTUDAR'
    if re.search(r'10[°º]',c): return 'ESTAGIÁRIO 10° TRIAGEM' if 'TRIAGEM' in c else 'ESTAGIÁRIO 10°'
    if re.search(r'9[°º]',c):  return 'ESTAGIÁRIO 9° TRIAGEM'  if 'TRIAGEM' in c else 'ESTAGIÁRIO 9°'
    en = normalize(est).strip()
    if en and not any(x in en.upper() for x in ['PSICODIAG','NÃO','MARCAR','SUPERVISÃO']):
        return 'ESTAGIÁRIO 9° TRIAGEM' if 'TRIAGEM' in en.upper() else 'ESTAGIÁRIO 9°'
    return 'LIVRE' if not en else 'OUTRO'

# Arquivo xlsx - ajuste o nome se necessário
ARQUIVO = 'Cópia de MAPA DE SALA 2026.1.xlsx'

conn = sqlite3.connect('mapa_salas.db')
conn.execute("DELETE FROM agendamentos")
conn.commit()

# Adicionar coluna dia_semana se não existir
try:
    conn.execute("ALTER TABLE agendamentos ADD COLUMN dia_semana TEXT DEFAULT 'SEGUNDA'")
    conn.commit()
except:
    pass

wb = openpyxl.load_workbook(ARQUIVO, data_only=True)
total_inserido = 0

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    dia = sheet_name.upper().strip()
    # Pular abas que não são dias da semana
    dias_validos = ['SEGUNDA','TERÇA','TERCA','QUARTA','QUINTA','SEXTA','SÁBADO','SABADO']
    if not any(d in dia for d in dias_validos):
        print(f"  Pulando aba: {sheet_name}")
        continue

    print(f"\nProcessando aba: {sheet_name}")
    rows = list(ws.iter_rows(values_only=True))

    cur_hor = None
    est_row = None
    inserido_dia = 0

    for row in rows[1:]:  # pular cabeçalho
        if not row or all(c is None or str(c).strip()=='' for c in row):
            continue
        first = str(row[0]).strip() if row[0] else ''
        if 'LEGENDA' in first.upper():
            break
        # Normalizar horário (07:00:00 -> 07:00)
        hor_match = re.match(r'^(\d{1,2}):(\d{2})', first)
        if hor_match:
            cur_hor = f"{int(hor_match.group(1)):02d}:{hor_match.group(2)}"
            est_row = row
        elif cur_hor and est_row is not None:
            pac_row = row
            for i, sala in enumerate(SALAS):
                est = normalize(est_row[i+1]) if i+1 < len(est_row) and est_row[i+1] else ''
                pac = normalize(pac_row[i+1]) if i+1 < len(pac_row) and pac_row[i+1] else ''
                if not est.strip() and not pac.strip():
                    continue
                cat = detect_cat(est, pac)
                tri = 1 if 'TRIAGEM' in (est+pac).upper() else 0
                conn.execute(
                    'INSERT INTO agendamentos(dia_semana,horario,sala,estagiario,paciente,categoria,triagem)'
                    ' VALUES(?,?,?,?,?,?,?)',
                    (dia, cur_hor, sala, est, pac, cat, tri)
                )
                inserido_dia += 1
            est_row = None

    conn.commit()
    print(f"  {inserido_dia} registros inseridos")
    total_inserido += inserido_dia

conn.close()
print(f"\n✅ TOTAL: {total_inserido} registros importados com sucesso!")
